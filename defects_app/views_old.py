import re
import json
import barcode
import os
import uuid
from urllib.parse import urlencode
from django.urls import reverse

from PIL import Image, ImageOps
from django.core.files.base import ContentFile

from django.db.models import Max

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponseBadRequest, Http404
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from .decorators import permission_required, station_session_required
from .permissions import (
    has_permission,
    is_manager_user,
    can_create_defects,
    can_edit_delete_defects,
    can_fix_aggregates,
    can_edit_aggregates,
    can_create_cars_and_print,
    can_view_reports_exports,
)



from django.contrib.auth import login, logout
from openpyxl import load_workbook
from django.db import IntegrityError
from io import BytesIO
from barcode.writer import SVGWriter

from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime, time, timedelta

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines

import base64
from openpyxl.drawing.image import Image as ExcelImage
from tempfile import NamedTemporaryFile

from calendar import monthrange
from datetime import date
from django.db.models import Count, Q, Min
from django.db.models import OuterRef, Subquery

from .services.history_service import save_avto_history, save_defect_history
from .services.photo_service import save_defect_photos
from .services.status_service import (
    car_passed_bestenevaya,
    car_is_on_sgp,
    car_is_on_snp,
    send_car_to_snp,
    send_car_to_sgp,
    approve_defects_for_sgp,
    add_snp_comment,
    mark_defect_fixed,
    mark_defect_verified,
)

from .models import (
    Avtomobili,
    Defekty,
    StatusAvto,
    Mesta,
    Smeny,
    IstoriyaIzmeneniyAvto,
    IstoriyaIzmeneniyDefektov,
    SnpDefectComment,
    SnpDefectOrder,
    PlanovyeVin,
    DailyProductionPlan,
    Otvetstvennye,
    DefectApprovalForSgp,
    VinPrefix,
    Modeli,
    Tipy,
    Oblasti,
    Greydy,
    DefectPhoto,
)

from .forms import (
    CarSearchForm,
    VIN_PREFIXES,
    DefectForm,
    CustomLoginForm,
    ManagerLoginForm,
    TelematikaForm,
    GlonassForm,
    BatareyaForm,
    PerednijDvigatelForm,
    ZadnijDvigatelForm,
)

def is_vh1_station(station_id=None, station_name=None):
    if station_id == 13:
        return True

    if not station_name:
        return False

    normalized_name = str(station_name).strip().lower().replace("-", " ")
    normalized_name = " ".join(normalized_name.split())
    return normalized_name in {"вх 1", "вх1", "входной контроль 1"}

def get_vh1_station():
    stations = Mesta.objects.all()
    for station in stations:
        if is_vh1_station(station_id=station.id, station_name=station.nazvanie):
            return station
    return None