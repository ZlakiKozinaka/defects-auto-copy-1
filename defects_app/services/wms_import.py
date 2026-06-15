import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from defects_app.models import (
    WmsBox,
    WmsBoxItem,
    WmsCase,
    WmsContainer,
    WmsLot,
    WmsPallet,
    WmsPalletPlacement,
    WmsOperation,
    Container,
)

REQUIRED_ALIASES = {
    "case_number": ["case no.", "case no", "case", "case number", "箱号", "箱编号"],
    "container_number": ["集装箱编号", "container no.", "container no", "container", "container number"],
    "box_number": ["box no.", "box no", "box", "box number", "包装箱号"],
    "part_number": ["sap part number", "part number", "part no.", "part no", "零件号", "物料号"],
    "quantity": ["total qty", "qty", "quantity", "数量", "总数量"],
}
OPTIONAL_ALIASES = {
    "chinese_name": ["chinese name", "中文名称", "零件名称", "名称"],
    "chinese_description": ["chinese description", "中文描述", "描述"],
    "english_name": ["english name", "英文名称", "en name"],
    "english_description": ["english description", "英文描述", "description"],
    "unit": ["unit", "单位"],
    "gross_weight": ["gross weight", "gross wt", "毛重"],
    "net_weight": ["net weight", "net wt", "净重"],
    "volume": ["volume", "体积", "cbm"],
}
LOT_ALIASES = ["lot", "lot no", "lot no.", "lot number", "批次", "批号"]


@dataclass
class WmsImportResult:
    lot: WmsLot
    containers_count: int
    cases_count: int
    boxes_count: int
    items_count: int
    skipped_rows: int


def normalize_header(value):
    if value is None:
        return ""

    text = str(value).strip().lower()

    # В китайских Excel иногда первая буква в "Сontainer number"
    # бывает кириллическая "с", а не английская "c".
    text = text.replace("с", "c")

    return re.sub(r"\s+", " ", text)


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def parse_decimal(value):
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            value = value.replace(",", ".").strip()
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def find_column(headers, aliases):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if normalize_header(header) in normalized_aliases:
            return index
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if any(alias in normalized for alias in normalized_aliases if alias):
            return index
    return None


def detect_header_row(sheet):
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), start=1):
        headers = list(row)
        matched = 0
        for aliases in REQUIRED_ALIASES.values():
            if find_column(headers, aliases) is not None:
                matched += 1
        if matched >= 4:
            return row_index, headers
    raise ValueError("Не удалось найти строку заголовков Excel. Проверьте наличие колонок Case No., 集装箱编号, Box No., SAP Part Number, Total QTY.")


def build_column_map(headers):
    column_map = {}
    for key, aliases in {**REQUIRED_ALIASES, **OPTIONAL_ALIASES}.items():
        column_map[key] = find_column(headers, aliases)
    return column_map


def guess_lot_number(source_file, explicit_lot_number=""):
    if explicit_lot_number:
        return explicit_lot_number.strip()

    stem = Path(source_file.name).stem.strip()

    # Например: D/MY19-26004-11, D-MY19-26004-11, MY19-26004-11
    match = re.search(
        r"(?:[A-Z]/)?[A-Z]{2}\d{2}[-/]\d{5}[-/]\d{2}",
        stem,
        flags=re.IGNORECASE,
    )

    if match:
        return match.group(0).replace("-", "/", 1).strip()

    return stem


def get_cell(row, column_map, key):
    index = column_map.get(key)
    if index is None or index >= len(row):
        return ""
    return normalize_value(row[index])


def raw_row_dict(headers, row):
    result = {}
    for index, header in enumerate(headers):
        if header is None or index >= len(row):
            continue
        value = row[index]
        if value is not None:
            result[str(header).strip()] = value.isoformat() if hasattr(value, "isoformat") else str(value)
    return result

def detect_lot209_header_row(sheet):
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True),
        start=1,
    ):
        headers = [normalize_header(value) for value in row]

        has_container = any("container" in value and "number" in value for value in headers)
        has_case = any("case number" in value for value in headers)
        has_part = any("part" in value and "number" in value for value in headers)
        has_quantity = any("quantity" in value for value in headers)

        if has_container and has_case and has_part and has_quantity:
            return row_index, list(row)

    return None, None


def build_lot209_column_map(headers):
    normalized_headers = [normalize_header(value) for value in headers]

    def find_by_contains(*parts):
        for index, header in enumerate(normalized_headers):
            if all(part in header for part in parts):
                return index
        return None

    return {
        "container_number": find_by_contains("container", "number"),
        "case_number": find_by_contains("case", "number"),
        "part_number": find_by_contains("part", "number"),
        "quantity": find_by_contains("quantity"),
        "chinese_name": find_column(headers, ["中文名称", "chinese"]),
        "english_name": find_column(headers, ["english", "en"]),
        "net_weight": find_column(headers, ["net weight"]),
        "gross_weight": find_column(headers, ["gross weight"]),
        "volume": find_column(headers, ["volume"]),
    }


def import_lot209_format(workbook, lot, source_file, user, file_bytes):
    skipped_rows = 0

    for sheet in workbook.worksheets:
        if sheet.title == "商务报关单":
            continue

        header_row_number, headers = detect_lot209_header_row(sheet)

        if not header_row_number:
            continue

        column_map = build_lot209_column_map(headers)

        required = ["container_number", "case_number", "part_number", "quantity"]
        missing = [key for key in required if column_map.get(key) is None]

        if missing:
            skipped_rows += sheet.max_row
            continue

        current_container_number = ""
        current_case_number = ""

        for excel_row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            container_number = get_cell(row, column_map, "container_number").upper()
            case_number = get_cell(row, column_map, "case_number")
            part_number = get_cell(row, column_map, "part_number").upper()
            quantity = parse_decimal(get_cell(row, column_map, "quantity"))

            if container_number:
                current_container_number = container_number

            if case_number:
                current_case_number = case_number

            container_number = current_container_number
            case_number = current_case_number

            if not (container_number and case_number and part_number):
                skipped_rows += 1
                continue

            box_number = f"BOX-{case_number}"

            existing_container = Container.objects.filter(number__iexact=container_number).first()

            container, _ = WmsContainer.objects.get_or_create(
                lot=lot,
                container_number=container_number,
                defaults={
                    "existing_container": existing_container,
                },
            )

            if existing_container and not container.existing_container_id:
                container.existing_container = existing_container
                container.save(update_fields=["existing_container"])

            case, _ = WmsCase.objects.get_or_create(
                container=container,
                case_number=case_number,
            )

            box, _ = WmsBox.objects.get_or_create(
                case=case,
                box_number=box_number,
            )

            item_defaults = {
                "row_number": excel_row_number,
                "chinese_name": get_cell(row, column_map, "chinese_name"),
                "chinese_description": "",
                "english_name": get_cell(row, column_map, "english_name"),
                "english_description": "",
                "quantity": quantity or Decimal("0"),
                "unit": "PCS",
                "gross_weight": parse_decimal(get_cell(row, column_map, "gross_weight")),
                "net_weight": parse_decimal(get_cell(row, column_map, "net_weight")),
                "volume": parse_decimal(get_cell(row, column_map, "volume")),
                "raw_data": {
                    "sheet": sheet.title,
                    "row": excel_row_number,
                    **raw_row_dict(headers, row),
                },
            }

            WmsBoxItem.objects.update_or_create(
                box=box,
                part_number=part_number,
                row_number=excel_row_number,
                defaults=item_defaults,
            )

    if not WmsContainer.objects.filter(lot=lot).exists():
        raise ValueError(
            "Не удалось распознать файл LOT209. "
            "Проверьте, что в листах есть колонки Container number, Case Number, Part Number, Quantity."
        )

    return skipped_rows

@transaction.atomic
def import_wms_lot_from_excel(source_file, lot_number, user, comment="", site=None):
    lot_number = guess_lot_number(source_file, lot_number)
    if not lot_number:
        raise ValueError("Укажите номер лота или загрузите файл с номером лота в имени.")

    file_bytes = source_file.read()
    source_file.seek(0)
    workbook = load_workbook(filename=source_file, data_only=True, read_only=True)

    sheet = None
    header_row_number = None
    headers = None
    import_format = None

    # 1. Сначала пробуем старый формат MY19.
    # Он обязан иметь Case No, Container, Box No, SAP Part Number, Total QTY.
    for candidate_sheet in workbook.worksheets:
        try:
            found_header_row_number, found_headers = detect_header_row(candidate_sheet)
            found_column_map = build_column_map(found_headers)

            standard_missing = [
                key for key in REQUIRED_ALIASES
                if found_column_map.get(key) is None
            ]

            if not standard_missing:
                sheet = candidate_sheet
                header_row_number = found_header_row_number
                headers = found_headers
                column_map = found_column_map
                import_format = "standard"
                break

        except ValueError:
            continue

    # 2. Если полноценный MY19 не найден — пробуем LOT209.
    if import_format is None:
        for candidate_sheet in workbook.worksheets:
            header_row_number_lot209, headers_lot209 = detect_lot209_header_row(candidate_sheet)

            if header_row_number_lot209:
                import_format = "lot209"
                break

    if import_format is None:
        raise ValueError(
            "Не удалось распознать Excel. "
            "Поддерживаются два формата: "
            "MY19 с колонками Case No., 集装箱编号, Box No., SAP Part Number, Total QTY "
            "или LOT209 с колонками Container number, Case Number, Part Number, Quantity."
        )

    if import_format == "standard":
        column_map = build_column_map(headers)
    else:
        column_map = {}

    if import_format == "standard":
        missing = [key for key in REQUIRED_ALIASES if column_map.get(key) is None]
        if missing:
            raise ValueError("Не найдены обязательные колонки: " + ", ".join(missing))

    lot, _ = WmsLot.objects.update_or_create(
        site=site,
        lot_number=lot_number,
        defaults={
            "display_name": Path(source_file.name).stem,
            "uploaded_by": getattr(user, "username", str(user)) if user else "",
            "uploaded_at": timezone.now(),
            "is_active": True,
            "comment": comment,
        }
    )
    lot.source_file.save(source_file.name, ContentFile(file_bytes), save=True)

    # Повторный импорт опасен, если по контейнерам этого лота уже есть поддоны/размещения.
    # В таком случае нельзя удалять структуру лота, иначе можно потерять складскую историю.
    has_pallets = WmsPallet.objects.filter(storage_unit__container__lot=lot).exists()
    has_placements = WmsPalletPlacement.objects.filter(pallet__storage_unit__container__lot=lot).exists()

    if has_pallets or has_placements:
        WmsOperation.objects.create(
            operation_type=WmsOperation.OP_REIMPORT_BLOCKED,
            lot=lot,
            message="Повторный импорт заблокирован: по лоту уже есть поддоны или складские размещения.",
            performed_by=getattr(user, "username", str(user)) if user else "",
            data={
                "source_file": source_file.name,
                "has_pallets": has_pallets,
                "has_placements": has_placements,
            },
        )
        raise ValueError(
            "Повторный импорт запрещён: по этому лоту уже есть поддоны или складские размещения. "
            "Чтобы не потерять историю склада, сначала проверьте лот вручную или создайте новый номер версии лота."
        )

    # Если лот уже загружался, но ещё не размещался на складе, можно безопасно пересоздать структуру.
    lot.containers.all().delete()

    skipped_rows = 0

    if import_format == "lot209":
        skipped_rows = import_lot209_format(
            workbook=workbook,
            lot=lot,
            source_file=source_file,
            user=user,
            file_bytes=file_bytes,
        )

        WmsOperation.objects.create(
            operation_type=WmsOperation.OP_IMPORT_LOT,
            lot=lot,
            message=f"Импортирован Excel-файл лота {lot.lot_number} в формате LOT209",
            performed_by=getattr(user, "username", str(user)) if user else "",
            data={
                "source_file": source_file.name,
                "format": "lot209",
                "containers_count": WmsContainer.objects.filter(lot=lot).count(),
                "cases_count": WmsCase.objects.filter(container__lot=lot).count(),
                "boxes_count": WmsBox.objects.filter(case__container__lot=lot).count(),
                "items_count": WmsBoxItem.objects.filter(box__case__container__lot=lot).count(),
                "skipped_rows": skipped_rows,
            },
        )

        return WmsImportResult(
            lot=lot,
            containers_count=WmsContainer.objects.filter(lot=lot).count(),
            cases_count=WmsCase.objects.filter(container__lot=lot).count(),
            boxes_count=WmsBox.objects.filter(case__container__lot=lot).count(),
            items_count=WmsBoxItem.objects.filter(box__case__container__lot=lot).count(),
            skipped_rows=skipped_rows,
        )

    for excel_row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
        start=header_row_number + 1,
    ):
        container_number = get_cell(row, column_map, "container_number").upper()
        case_number = get_cell(row, column_map, "case_number")
        box_number = get_cell(row, column_map, "box_number")
        part_number = get_cell(row, column_map, "part_number").upper()
        quantity = parse_decimal(get_cell(row, column_map, "quantity"))

        if not (container_number and case_number and box_number and part_number):
            skipped_rows += 1
            continue

        existing_container = Container.objects.filter(number__iexact=container_number).first()
        container, _ = WmsContainer.objects.get_or_create(
            lot=lot,
            container_number=container_number,
            defaults={"existing_container": existing_container},
        )
        if existing_container and not container.existing_container_id:
            container.existing_container = existing_container
            container.save(update_fields=["existing_container"])

        case, _ = WmsCase.objects.get_or_create(container=container, case_number=case_number)
        box, _ = WmsBox.objects.get_or_create(case=case, box_number=box_number)

        item_defaults = {
            "row_number": excel_row_number,
            "chinese_name": get_cell(row, column_map, "chinese_name"),
            "chinese_description": get_cell(row, column_map, "chinese_description"),
            "english_name": get_cell(row, column_map, "english_name"),
            "english_description": get_cell(row, column_map, "english_description"),
            "quantity": quantity or Decimal("0"),
            "unit": get_cell(row, column_map, "unit"),
            "gross_weight": parse_decimal(get_cell(row, column_map, "gross_weight")),
            "net_weight": parse_decimal(get_cell(row, column_map, "net_weight")),
            "volume": parse_decimal(get_cell(row, column_map, "volume")),
            "raw_data": raw_row_dict(headers, row),
        }
        WmsBoxItem.objects.update_or_create(
            box=box,
            part_number=part_number,
            row_number=excel_row_number,
            defaults=item_defaults,
        )

    WmsOperation.objects.create(
        operation_type=WmsOperation.OP_IMPORT_LOT,
        lot=lot,
        message=f"Импортирован Excel-файл лота {lot.lot_number}",
        performed_by=getattr(user, "username", str(user)) if user else "",
        data={
            "source_file": source_file.name,
            "containers_count": WmsContainer.objects.filter(lot=lot).count(),
            "cases_count": WmsCase.objects.filter(container__lot=lot).count(),
            "boxes_count": WmsBox.objects.filter(case__container__lot=lot).count(),
            "items_count": WmsBoxItem.objects.filter(box__case__container__lot=lot).count(),
            "skipped_rows": skipped_rows,
        },
    )

    return WmsImportResult(
        lot=lot,
        containers_count=WmsContainer.objects.filter(lot=lot).count(),
        cases_count=WmsCase.objects.filter(container__lot=lot).count(),
        boxes_count=WmsBox.objects.filter(case__container__lot=lot).count(),
        items_count=WmsBoxItem.objects.filter(box__case__container__lot=lot).count(),
        skipped_rows=skipped_rows,
    )