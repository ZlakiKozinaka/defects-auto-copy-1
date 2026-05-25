import re
import json
import barcode
import os
import uuid
from urllib.parse import urlencode
from django.urls import reverse

from PIL import Image, ImageOps
from django.core.files.base import ContentFile

from django.db.models import Max

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponseBadRequest, Http404
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from openpyxl import load_workbook
from django.db import IntegrityError
from io import BytesIO
from barcode.writer import SVGWriter

from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime, time, timedelta

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines

import base64
from openpyxl.drawing.image import Image as ExcelImage
from tempfile import NamedTemporaryFile

from calendar import monthrange
from datetime import date
from django.db.models import Count, Q, Min
from django.db.models import OuterRef, Subquery

from .models import (
    Avtomobili,
    Defekty,
    StatusAvto,
    Mesta,
    Smeny,
    IstoriyaIzmeneniyAvto,
    IstoriyaIzmeneniyDefektov,
    SnpDefectComment,
    SnpDefectOrder,
    PlanovyeVin,
    DailyProductionPlan,
    Otvetstvennye,
    DefectApprovalForSgp,
    VinPrefix,
    Modeli,
    Tipy,
    Oblasti,
    Greydy,
    DefectPhoto,
)

from .forms import (
    CarSearchForm,
    VIN_PREFIXES,
    DefectForm,
    CustomLoginForm,
    ManagerLoginForm,
    TelematikaForm,
    GlonassForm,
    BatareyaForm,
    PerednijDvigatelForm,
    ZadnijDvigatelForm,
)


def user_in_group(user, group_name):
    return user.groups.filter(name=group_name).exists()


def is_manager_user(user):
    return user.groups.filter(name="Начальники").exists() or user.is_superuser

def has_any_group(user, group_names):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=group_names).exists()


# --- Роли ОТК ---
def is_otk_worker(user):
    return user_in_group(user, "Работник ОТК")


def is_otk_manager(user):
    return user_in_group(user, "Начальник ОТК")


# --- Роли Агрегатов ---
def is_agg_worker(user):
    return user_in_group(user, "Работник Агрегатов")


def is_agg_manager(user):
    return user_in_group(user, "Начальник Агрегатов")


# --- Роли Логистики ---
def is_log_worker(user):
    return user_in_group(user, "Работник Логистики")


def is_log_manager(user):
    return user_in_group(user, "Начальник Логистики")


# --- Планирование ---
def is_planning_user(user):
    return user_in_group(user, "Отдел планирования производства")


# --- Агрегированные проверки прав ---
def can_create_defects(user):
    # Работник ОТК + Начальник ОТК + глобальные начальники
    return is_manager_user(user) or is_otk_worker(user) or is_otk_manager(user)


def can_edit_delete_defects(user):
    # Только начальники ОТК и глобальные начальники
    return is_manager_user(user) or is_otk_manager(user)


def can_fix_aggregates(user):
    # Работник Агрегатов + Начальник Агрегатов + глобальные начальники
    return is_manager_user(user) or is_agg_worker(user) or is_agg_manager(user)


def can_edit_aggregates(user):
    # Только начальники Агрегатов и глобальные начальники
    return is_manager_user(user) or is_agg_manager(user)


def can_create_cars_and_print(user):
    # Работник Логистики + Начальник Логистики + глобальные начальники
    return is_manager_user(user) or is_log_worker(user) or is_log_manager(user)


def can_view_reports_exports(user):
    # Все начальники направлений + планирование + глобальные начальники
    return (
        is_manager_user(user)
        or is_otk_manager(user)
        or is_agg_manager(user)
        or is_log_manager(user)
        or is_planning_user(user)
    )

@login_required
def department_hub_view(request):
    departments = get_user_departments(request.user)

    if not departments:
        logout(request)
        messages.error(request, "Для вашей учетной записи не назначена ни одна роль.")
        return redirect("login")

    # если только одна принадлежность — сразу туда
    if len(departments) == 1:
        only_dep = list(departments)[0]
        return redirect("department_section", section=only_dep)

    return render(request, "defects_app/department_hub.html", {
        "departments": departments,
    })

def user_has_department_access(user, department_key):
    departments = get_user_departments(user)
    return department_key in departments

def render_access_denied(request, message=None):
    return render(request, "403.html", {
        "access_denied_message": message or "Проверьте, что выбран правильный отдел и роль пользователя.",
    }, status=403)

@login_required
def department_section_view(request, section):
    if not user_has_department_access(request.user, section):
        return render_access_denied(request, "Вы не имеете доступа к этому отделу. Проверьте, что выбрали правильный отдел.")

    section_titles = {
        "otk": "ОТК",
        "aggregates": "Агрегаты",
        "logistics": "Логистика",
        "planning": "Планирование производства",
    }

    return render(request, "defects_app/department_section.html", {
        "section": section,
        "section_title": section_titles.get(section, section),
        "is_global_manager": is_manager_user(request.user),
    })

def get_station_redirect_name(station_id):
    station_map = {
        1: "home",                 # Бестеневая
        2: "okline",
        13: "vh1",
        3: "telematika_glonass",
        4: "telematika_glonass",
        5: "agregaty",
        6: "agregaty",
        7: "agregaty",
        11: "dovodka",
        12: "create_car",
    }

    if station_id in [8, 9, 10]:
        return "quality"

    return station_map.get(station_id, "home")


def is_vh1_station(station_id=None, station_name=None):
    if station_id == 13:
        return True

    if not station_name:
        return False

    normalized_name = str(station_name).strip().lower().replace("-", " ")
    normalized_name = " ".join(normalized_name.split())
    return normalized_name in {"вх 1", "вх1", "входной контроль 1"}


def get_vh1_station():
    stations = Mesta.objects.all()
    for station in stations:
        if is_vh1_station(station_id=station.id, station_name=station.nazvanie):
            return station
    return None

def get_user_departments(user):
    departments = set()

    if is_manager_user(user):
        departments.update(["otk", "aggregates", "logistics", "planning"])

    if is_otk_worker(user) or is_otk_manager(user):
        departments.add("otk")

    if is_agg_worker(user) or is_agg_manager(user):
        departments.add("aggregates")

    if is_log_worker(user) or is_log_manager(user):
        departments.add("logistics")

    if is_planning_user(user):
        departments.add("planning")

    return departments

def require_station_session(request):
    station_id = request.session.get("station_id")
    shift_id = request.session.get("shift_id")

    if not request.user.is_authenticated:
        return None, redirect("login")

    if not station_id or not shift_id:
        logout(request)
        messages.info(request, "Сессия истекла. Войдите снова.")
        return None, redirect("login")

    return {
        "station_id": station_id,
        "shift_id": shift_id,
        "station_name": request.session.get("station_name"),
        "shift_name": request.session.get("shift_name"),
    }, None



def csrf_failure_view(request, reason=""):
    return render(request, "defects_app/csrf_error.html", {
        "reason": reason,
    }, status=403)

def manager_required_view(request):
    if not request.user.is_authenticated:
        return redirect("login")

    if not is_manager_user(request.user):
        return HttpResponseForbidden("Доступ только для начальников.")

    return None


def parse_glonass_data(glonass_text):
    if not glonass_text:
        return None, None, None

    text = glonass_text.strip()

    sn_match = re.search(r"SN\s*:\s*([^;]+)", text, re.IGNORECASE)
    imei_match = re.search(r"IMEI\s*:\s*([^;]+)", text, re.IGNORECASE)
    iccid_match = re.search(r"ICCID\s*:\s*([^;]+)", text, re.IGNORECASE)

    sn = sn_match.group(1).strip() if sn_match else None
    imei = imei_match.group(1).strip() if imei_match else None
    iccid = iccid_match.group(1).strip() if iccid_match else None

    return sn, imei, iccid

def compress_defect_photo(uploaded_file, defect_id):
    max_size = 1280
    quality = 70

    image = Image.open(uploaded_file)

    if image.mode in ("RGBA", "P"):
        image = image.convert("RGB")
    else:
        image = image.convert("RGB")

    image.thumbnail((max_size, max_size))

    buffer = BytesIO()
    image.save(buffer, format="WEBP", quality=quality, method=6)

    file_name = f"defect_{defect_id}_{uuid.uuid4().hex}.webp"

    return ContentFile(buffer.getvalue(), name=file_name)

def save_defect_photos(defect, request):
    photos = request.FILES.getlist("photos")

    for uploaded_file in photos:
        try:
            image = Image.open(uploaded_file)
            image = ImageOps.exif_transpose(image)

            if image.mode not in ("RGB",):
                image = image.convert("RGB")

            max_size = (1280, 1280)
            image.thumbnail(max_size, Image.Resampling.LANCZOS)

            buffer = BytesIO()
            image.save(
                buffer,
                format="WEBP",
                quality=70,
                method=6
            )

            file_name = f"{uuid.uuid4().hex}.webp"

            photo = DefectPhoto(
                defect=defect,
                avto=defect.avto,
                original_name=uploaded_file.name,
                uploaded_by=request.user.username,
            )

            photo.image.save(
                file_name,
                ContentFile(buffer.getvalue()),
                save=True
            )

            photo.file_size = photo.image.size
            photo.save(update_fields=["file_size"])

        except Exception as e:
            messages.error(
                request,
                f"Не удалось обработать фото: {uploaded_file.name}. Ошибка: {e}"
            )

def custom_login_view(request):
    if request.user.is_authenticated:
        next_url = request.GET.get("next")
        if next_url:
            return redirect(next_url)
        return redirect("department_hub")

    if request.method == "POST":
        form = CustomLoginForm(request, data=request.POST)

        if form.is_valid():
            user = form.get_user()
            shift = form.cleaned_data["shift"]
            department = form.cleaned_data["department"]

            # ВАЖНО: проверяем доступ к выбранному отделу ДО login()
            if not user_has_department_access(user, department):
                return render_access_denied(request, "Вы не имеете доступа к выбранному отделу. Проверьте, что выбрали правильный отдел.")


            # Только после проверки прав — логиним
            login(request, user)

            # Смена
            request.session["shift_id"] = shift.id
            request.session["shift_name"] = shift.nazvanie

            # Выбранный отдел
            request.session["department_key"] = department

            department_titles = {
                "otk": "ОТК",
                "aggregates": "Агрегаты",
                "logistics": "Логистика",
                "planning": "Планирование производства",
            }
            request.session["department_name"] = department_titles.get(department, department)

            # Очищаем station-контекст — станция теперь выбирается кнопкой в разделе
            request.session.pop("station_id", None)
            request.session.pop("station_name", None)

            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)

            # Сразу в выбранный раздел
            return redirect("department_section", section=department)
    else:
        form = CustomLoginForm()

    return render(request, "registration/login.html", {"form": form})

def save_avto_history(avto, field_name, old_value, new_value, username):
    old_str = "" if old_value is None else str(old_value)
    new_str = "" if new_value is None else str(new_value)

    # Историю пишем только если поле уже было заполнено раньше
    if old_str and old_str != new_str:
        IstoriyaIzmeneniyAvto.objects.create(
            avto=avto,
            pole=field_name,
            staroe_znachenie=old_str,
            novoe_znachenie=new_str,
            kto_izmenil=username,
        )


def save_defect_history(defect, field_name, old_value, new_value, username):
    old_str = "" if old_value is None else str(old_value)
    new_str = "" if new_value is None else str(new_value)

    if old_str != new_str:
        IstoriyaIzmeneniyDefektov.objects.create(
            defekt=defect,
            pole=field_name,
            staroe_znachenie=old_str,
            novoe_znachenie=new_str,
            kto_izmenil=username,
        )

def car_passed_bestenevaya(car):
    return bool(car.proshla_bestenevaya and car.data_prohoda_bestenevaya)

@login_required
def home(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]
    is_manager = can_edit_delete_defects(request.user)

    if not can_create_defects(request.user):
        return HttpResponseForbidden("У вас нет доступа к рабочей вкладке ОТК.")

    form = CarSearchForm()
    car = None
    defects = []

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })
        defects = Defekty.objects.filter(avto=car).order_by("-data")

    if request.method == "POST":
        action = request.POST.get("action")        

        if action == "save" and not can_create_cars_and_print(request.user):
            messages.error(request, "Создание машины доступно только логистике и начальникам.")
            return render(request, "defects_app/home.html", {
                "form": form,
                "car": car,
                "defects": defects,
                "vin_prefixes": VIN_PREFIXES,
                "is_manager": is_manager,
            })

        current_car_id = request.POST.get("current_car_id")
        if current_car_id and action in ["find", "save"]:
            current_car = get_object_or_404(Avtomobili, id=current_car_id)

            if not current_car.proshla_bestenevaya:
                messages.error(
                    request,
                    "Сначала завершите текущую машину: нажмите «Машина прошла станцию Бестеневая»."
                )

                defects = Defekty.objects.filter(avto=current_car).order_by("-data")
                form = CarSearchForm(initial={
                    "vin": current_car.vin,
                    "model": current_car.model
                })

                return render(request, "defects_app/home.html", {
                    "form": form,
                    "car": current_car,
                    "defects": defects,
                    "vin_prefixes": VIN_PREFIXES,
                    "is_manager": is_manager,
                })

        form = CarSearchForm(request.POST)

        if form.is_valid():
            vin = form.cleaned_data["vin"]
            model = form.cleaned_data["model"]

            existing_car = Avtomobili.objects.filter(vin=vin).first()

            if action == "find":
                if existing_car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/?car_id={existing_car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")
                    return render(request, "defects_app/home.html", {
                        "form": form,
                        "car": None,
                        "defects": [],
                        "vin_prefixes": VIN_PREFIXES,
                        "is_manager": is_manager,
                    })

            elif action == "save":
                if existing_car:
                    messages.info(request, "Машина с таким VIN уже существует. Открыта существующая запись.")
                    return redirect(f"/?car_id={existing_car.id}")
                else:
                    car = Avtomobili.objects.create(
                        vin=vin,
                        model=model,
                        kto_sozdal=request.user.username,
                        data_sozdaniya=timezone.now(),
                    )
                    messages.success(request, "Новая машина успешно создана.")
                    return redirect(f"/?car_id={car.id}")

    return render(request, "defects_app/home.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager,
    })


@login_required
def create_defect(request, car_id):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    if not can_create_defects(request.user):
        return HttpResponseForbidden("У вас нет прав на создание дефектов.")

    station_id = session_data["station_id"]
    shift_id = session_data["shift_id"]

    car = get_object_or_404(Avtomobili, id=car_id)
    from_page = request.GET.get("from_page", "") if request.method == "GET" else request.POST.get("from_page", "")

    if from_page == "okline":
        last_status = StatusAvto.objects.filter(avto=car).order_by("-data_statusa").first()
        if last_status and last_status.status == "СГП":
            messages.error(request, "Для машины, уже переданной на СГП, нельзя создавать новые дефекты.")
            return redirect(f"/okline/?car_id={car.id}")

    if from_page == "okline" and station_id != 2:
        return HttpResponseForbidden("У вас нет доступа к созданию дефектов с OKLine.")

    if from_page == "quality" and station_id not in [8, 9, 10]:
        return HttpResponseForbidden("У вас нет доступа к созданию дефектов со станции качества.")

    station_name = session_data.get("station_name")

    if from_page == "vh1" and not is_vh1_station(station_id=station_id, station_name=station_name):
        return HttpResponseForbidden("У вас нет доступа к созданию дефектов с ВХ1.")

    if from_page == "dovodka" and station_id != 11:
        return HttpResponseForbidden("У вас нет доступа к созданию дефектов с Доводки.")

    if request.method == "POST":
        form = DefectForm(request.POST, request.FILES)

        if form.is_valid():
            defect = form.save(commit=False)
            defect.avto = car
            defect.kto_sozdal = request.user.username

            auto_mesto = Mesta.objects.filter(id=station_id).first()
            auto_smena = Smeny.objects.filter(id=shift_id).first()

            if not auto_mesto or not auto_smena:
                logout(request)
                messages.error(request, "Не найдены станция или смена. Войдите заново.")
                return redirect("login")

            defect.mesto = auto_mesto
            defect.smena = auto_smena
            defect.save()
            save_defect_photos(defect, request)

            messages.success(request, "Дефект успешно добавлен.")

            if from_page == "okline":
                return redirect(f"/okline/?car_id={car.id}")
            elif from_page == "quality":
                return redirect(f"/quality/?car_id={car.id}")
            elif from_page == "vh1":
                return redirect(f"/vh1/?car_id={car.id}")
            elif from_page == "dovodka":
                return redirect(f"/dovodka/?car_id={car.id}")
            return redirect(f"/?car_id={car.id}")
    else:
        form = DefectForm()

    return render(request, "defects_app/create_defect.html", {
        "form": form,
        "car": car,
        "from_page": from_page,
    })


@login_required
def check_vin(request):
    vin = request.GET.get("vin", "").strip().upper()

    if len(vin) != 17:
        return JsonResponse({"found": False})

    car = Avtomobili.objects.filter(
        vin=vin,
        created_on_station_1=True
    ).first()

    if car:
        return JsonResponse({
            "found": True,
            "car_id": car.id
        })

    return JsonResponse({"found": False})


@login_required
def edit_defect(request, defect_id):
    if not can_edit_delete_defects(request.user):
        return HttpResponseForbidden("У вас нет прав на редактирование дефектов.")

    defect = get_object_or_404(Defekty, id=defect_id)
    car = defect.avto
    from_page = request.GET.get("from_page", "") if request.method == "GET" else request.POST.get("from_page", "")

    if request.method == "POST":
        # Берем оригинальные значения из базы ДО валидации формы
        original_defect = Defekty.objects.get(id=defect.id)

        form = DefectForm(request.POST, request.FILES, instance=defect)
        if form.is_valid():
            updated_defect = form.save(commit=False)
            updated_defect.avto = car
            updated_defect.kto_sozdal = original_defect.kto_sozdal
            updated_defect.data = original_defect.data
            updated_defect.mesto = original_defect.mesto
            updated_defect.smena = original_defect.smena
            updated_defect.save()
            save_defect_photos(updated_defect, request)

            save_defect_history(defect, "smena", original_defect.smena, updated_defect.smena, request.user.username)
            save_defect_history(defect, "mesto", original_defect.mesto, updated_defect.mesto, request.user.username)
            save_defect_history(defect, "tip", original_defect.tip, updated_defect.tip, request.user.username)
            save_defect_history(defect, "oblast", original_defect.oblast, updated_defect.oblast, request.user.username)
            save_defect_history(defect, "greyd", original_defect.greyd, updated_defect.greyd, request.user.username)
            save_defect_history(defect, "otvetstvennyj", original_defect.otvetstvennyj, updated_defect.otvetstvennyj, request.user.username)
            save_defect_history(defect, "stanciya", original_defect.stanciya, updated_defect.stanciya, request.user.username)
            save_defect_history(defect, "kommentarij", original_defect.kommentarij, updated_defect.kommentarij, request.user.username)

            messages.success(request, "Изменения по дефекту сохранены.")

            if from_page == "quality":
                return redirect(f"/quality/?car_id={car.id}")

            if from_page == "vh1":
                return redirect(f"/vh1/?car_id={car.id}")

            if from_page == "okline":
                return redirect(f"/okline/?car_id={car.id}")
            
            if from_page == "dovodka":
                return redirect(f"/dovodka/?car_id={car.id}")

            return redirect(f"/?car_id={car.id}")
    else:
        form = DefectForm(instance=defect)

    return render(request, "defects_app/edit_defect.html", {
        "form": form,
        "car": car,
        "defect": defect,
        "from_page": from_page,
    })


@login_required
def delete_defect(request, defect_id):
    if not can_edit_delete_defects(request.user):
        return HttpResponseForbidden("У вас нет прав на удаление дефектов.")

    defect = get_object_or_404(Defekty, id=defect_id)
    car = defect.avto
    from_page = request.GET.get("from_page", "") if request.method == "GET" else request.POST.get("from_page", "")

    if request.method == "POST":
        defect.delete()
        messages.success(request, "Дефект удален.")

        if from_page == "quality":
            return redirect(f"/quality/?car_id={car.id}")

        if from_page == "okline":
            return redirect(f"/okline/?car_id={car.id}")

        if from_page == "vh1":
            return redirect(f"/vh1/?car_id={car.id}")

        return redirect(f"/?car_id={car.id}")

    return render(request, "defects_app/delete_defect.html", {
        "defect": defect,
        "car": car,
        "from_page": from_page,
    })


@login_required
def okline_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    is_manager = can_edit_delete_defects(request.user)

    station_id = session_data["station_id"]
    if not can_create_defects(request.user):
        return HttpResponseForbidden("У вас нет доступа к вкладке OKLine.")

    form = CarSearchForm()
    car = None
    defects = []
    is_sgp_locked = False
    all_verified = False
    snp_defects = []
    sgp_problem_defects = []

    car_id = request.GET.get("car_id")

    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)

        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        defects = Defekty.objects.filter(avto=car).prefetch_related(
            "snp_comments"
        ).order_by("-data")

        snp_defects = Defekty.objects.filter(
            avto=car,
            proveren=False
        ).order_by("-data")

        sgp_problem_defects = Defekty.objects.filter(
            avto=car
        ).filter(
            Q(ustraneno=False) | Q(proveren=False)
        ).filter(
            sgp_approval__isnull=True
        ).order_by("-data")

        last_status = StatusAvto.objects.filter(
            avto=car
        ).order_by("-data_statusa").first()

        if last_status and last_status.status == "СГП":
            is_sgp_locked = True

        all_verified = not defects.exists() or all(
            defect.ustraneno and defect.proveren for defect in defects
        )

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    return redirect(f"/okline/?car_id={car.id}")

                messages.error(request, "Машина с таким VIN не найдена.")

            return render(request, "defects_app/okline.html", {
                "form": form,
                "car": None,
                "defects": [],
                "vin_prefixes": VIN_PREFIXES,
                "is_sgp_locked": False,
                "all_verified": False,
                "is_manager": is_manager,
                "snp_defects": [],
                "sgp_problem_defects": [],
            })

        elif action == "save_checks":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            defects = Defekty.objects.filter(avto=car).order_by("-data")
            verified_ids = request.POST.getlist("verified_defects")

            verified_count = 0

            for defect in defects:
                if defect.ustraneno and not defect.proveren and str(defect.id) in verified_ids:
                    defect.proveren = True
                    defect.kto_razreshil = request.user.username
                    defect.data_razresheniya = timezone.now()
                    defect.save()
                    verified_count += 1

            if verified_count > 0:
                messages.success(request, f"Подтверждено дефектов: {verified_count}.")
            else:
                messages.info(request, "Новых отметок проверки не было.")

            return redirect(f"/okline/?car_id={car.id}")

        elif action == "send_snp":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)
            if not car_passed_bestenevaya(car):
                messages.error(
                    request,
                    "Нельзя отправить машину на СНП. Сначала машина должна пройти Бестеневую."
                )
                return redirect(f"/okline/?car_id={car.id}")

            defects = Defekty.objects.filter(avto=car)

            last_status = StatusAvto.objects.filter(
                avto=car
            ).order_by("-data_statusa").first()

            if last_status and last_status.status == "СГП":
                messages.error(request, "Машина уже находится в СГП. На СНП её отправить нельзя.")
                return redirect(f"/okline/?car_id={car.id}")

            all_verified_now = not defects.exists() or all(
                defect.ustraneno and defect.proveren
                for defect in defects
            )

            if last_status and last_status.status == "СНП":
                messages.info(request, "Машина уже находится на СНП.")
                return redirect(f"/okline/?car_id={car.id}")

            if all_verified_now:
                messages.error(request, "У машины нет неподтвержденных дефектов. На СНП её отправлять нельзя, только на СГП.")
                return redirect(f"/okline/?car_id={car.id}")

            snp_defects = Defekty.objects.filter(
                avto=car,
                proveren=False
            ).order_by("-data")

            if snp_defects.exists():
                for defect in snp_defects:
                    comment_text = request.POST.get(f"snp_comment_{defect.id}", "").strip()

                    if not comment_text:
                        messages.error(request, "Для каждого неподтвержденного дефекта нужно указать причину отправки на СНП.")
                        return redirect(f"/okline/?car_id={car.id}")

                for defect in snp_defects:
                    comment_text = request.POST.get(f"snp_comment_{defect.id}", "").strip()

                    SnpDefectComment.objects.create(
                        defect=defect,
                        avto=car,
                        comment=comment_text,
                        kto_sozdal=request.user.username,
                    )

            reset_count = 0

            for defect in defects:
                if defect.ustraneno and not defect.proveren:
                    defect.ustraneno = False
                    defect.kto_ustranil = None
                    defect.data_ustraneniya = None
                    defect.save()
                    reset_count += 1

            StatusAvto.objects.create(
                avto=car,
                status="СНП",
                kto_izmenil=request.user.username
            )

            if reset_count > 0:
                messages.success(request, f"Машина передана на СНП. Сброшено отметок устранения: {reset_count}.")
            else:
                messages.success(request, "Машина передана на СНП.")

            return redirect(f"/okline/?car_id={car.id}")

        elif action == "send_sgp":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)
            if not car_passed_bestenevaya(car):
                messages.error(
                    request,
                    "Нельзя отправить машину на СГП. Сначала машина должна пройти Бестеневую."
                )
                return redirect(f"/okline/?car_id={car.id}")

            last_status = StatusAvto.objects.filter(
                avto=car
            ).order_by("-data_statusa").first()

            if last_status and last_status.status == "СГП":
                messages.error(request, "Машина уже находится в СГП.")
                return redirect(f"/okline/?car_id={car.id}")

            problem_defects = Defekty.objects.filter(
                avto=car
            ).filter(
                Q(ustraneno=False) | Q(proveren=False)
            ).filter(
                sgp_approval__isnull=True
            ).order_by("-data")

            if problem_defects.exists():
                if not can_edit_delete_defects(request.user):
                    messages.error(
                        request,
                        "Ваших полномочий не хватает, чтобы отправить машину на СГП при неустраненных и(или) непроверенных дефектах. Пожалуйста, обратитесь к старшим."
                    )
                    return redirect(f"/okline/?car_id={car.id}")

                messages.error(
                    request,
                    "У машины есть неустраненные и(или) непроверенные дефекты. Для отправки на СГП необходимо согласование."
                )
                return redirect(f"/okline/?car_id={car.id}")

            StatusAvto.objects.create(
                avto=car,
                status="СГП",
                kto_izmenil=request.user.username
            )

            messages.success(request, "Машина передана на СГП.")
            return redirect("print_sgp_report", car_id=car.id)

        elif action == "approve_and_send_sgp":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            if not car_passed_bestenevaya(car):
                messages.error(
                    request,
                    "Нельзя отправить машину на СГП. Сначала машина должна пройти Бестеневую."
                )
                return redirect(f"/okline/?car_id={car.id}")

            if not can_edit_delete_defects(request.user):
                messages.error(
                    request,
                    "Ваших полномочий не хватает для согласования дефектов перед отправкой на СГП."
                )
                return redirect(f"/okline/?car_id={car.id}")

            last_status = StatusAvto.objects.filter(
                avto=car
            ).order_by("-data_statusa").first()

            if last_status and last_status.status == "СГП":
                messages.error(request, "Машина уже находится в СГП.")
                return redirect(f"/okline/?car_id={car.id}")

            comment = request.POST.get("approval_comment", "").strip()

            problem_defects = Defekty.objects.filter(
                avto=car
            ).filter(
                Q(ustraneno=False) | Q(proveren=False)
            ).filter(
                sgp_approval__isnull=True
            ).order_by("-data")

            if not problem_defects.exists():
                StatusAvto.objects.create(
                    avto=car,
                    status="СГП",
                    kto_izmenil=request.user.username
                )

                messages.success(request, "Машина передана на СГП.")
                return redirect("print_sgp_report", car_id=car.id)

            approved_count = problem_defects.count()

            for defect in problem_defects:
                DefectApprovalForSgp.objects.create(
                    defect=defect,
                    avto=car,
                    approved_by=request.user.username,
                    comment=comment
                )

            StatusAvto.objects.create(
                avto=car,
                status="СГП",
                kto_izmenil=request.user.username
            )

            messages.success(
                request,
                f"Согласовано дефектов: {approved_count}. Машина передана на СГП."
            )
            return redirect("print_sgp_report", car_id=car.id)

    return render(request, "defects_app/okline.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "vin_prefixes": VIN_PREFIXES,
        "is_sgp_locked": is_sgp_locked,
        "all_verified": all_verified,
        "is_manager": can_edit_delete_defects(request.user),
        "snp_defects": snp_defects,
        "sgp_problem_defects": sgp_problem_defects,
    })


@login_required
def complete_bestenevaya(request, car_id):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    is_manager = can_edit_delete_defects(request.user)
    station_id = session_data["station_id"]

    if not (can_create_cars_and_print(request.user) or can_create_defects(request.user)):
        return HttpResponseForbidden("У вас нет прав отмечать прохождение Бестеневой.")

    car = get_object_or_404(Avtomobili, id=car_id)

    if request.method == "POST":
        car.proshla_bestenevaya = True
        car.data_prohoda_bestenevaya = timezone.now()
        car.save(update_fields=["proshla_bestenevaya", "data_prohoda_bestenevaya"])

        messages.success(request, "Машина отмечена как прошедшая станцию Бестеневая.")
        return redirect(f"/?car_id={car.id}")

    return redirect(f"/?car_id={car.id}")


@login_required
def telematika_view(request):
    is_manager = is_manager_user(request.user)
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]
    if car.telematika and not can_edit_aggregates(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    

    form = CarSearchForm()
    telematika_form = TelematikaForm()
    car = None
    locked = False

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        if car.telematika:
            telematika_form = TelematikaForm(initial={
                "telematika": car.telematika
            })

            if not is_manager:
                locked = True

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)
            telematika_form = TelematikaForm()

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/telematika/?car_id={car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")

        elif action == "bind":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            if car.telematika and not is_manager:
                messages.error(request, "Телематика уже привязана. Изменение доступно только начальнику.")
                return redirect(f"/telematika/?car_id={car.id}")

            form = CarSearchForm(initial={
                "vin": car.vin,
                "model": car.model
            })
            telematika_form = TelematikaForm(request.POST)

            old_telematika = car.telematika

            if telematika_form.is_valid():
                new_telematika = telematika_form.cleaned_data["telematika"]

                save_avto_history(car, "telematika", old_telematika, new_telematika, request.user.username)

                car.telematika = new_telematika
                
                if not old_telematika:
                    car.privyazal_telematiku = request.user.username
                    car.data_privyazki_telematiki = timezone.now()
                    car.save(update_fields=[
                        "telematika",
                        "privyazal_telematiku",
                        "data_privyazki_telematiki",
                    ])
                else:
                    car.save(update_fields=["telematika"])

                if old_telematika and is_manager:
                    messages.success(request, "Телематика успешно изменена.")
                else:
                    messages.success(request, "Телематика успешно привязана.")

                return redirect(f"/telematika/?car_id={car.id}")

    return render(request, "defects_app/telematika.html", {
        "form": form,
        "telematika_form": telematika_form,
        "car": car,
        "vin_prefixes": VIN_PREFIXES,
        "locked": locked,
        "is_manager": is_manager,
    })



@login_required
def glonass_view(request):
    is_manager = is_manager_user(request.user)
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]
    if not can_fix_aggregates(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    

    form = CarSearchForm()
    glonass_form = GlonassForm()
    car = None
    locked = False

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        if car.glonass:
            glonass_form = GlonassForm(initial={
                "glonass": car.glonass
            })

            if not is_manager:
                locked = True

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)
            glonass_form = GlonassForm()

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/glonass/?car_id={car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")

        elif action == "bind":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            if car.glonass and not can_edit_aggregates(request.user):
                messages.error(request, "ГЛОНАСС уже привязан. Изменение доступно только начальнику.")
                return redirect(f"/glonass/?car_id={car.id}")

            form = CarSearchForm(initial={
                "vin": car.vin,
                "model": car.model
            })
            glonass_form = GlonassForm(request.POST)

            old_glonass = car.glonass
            old_sn = car.glonass_sn
            old_imei = car.glonass_imei
            old_iccid = car.glonass_iccid

            if glonass_form.is_valid():
                glonass_text = glonass_form.cleaned_data["glonass"]
                sn, imei, iccid = parse_glonass_data(glonass_text)

                save_avto_history(car, "glonass", old_glonass, glonass_text, request.user.username)
                save_avto_history(car, "glonass_sn", old_sn, sn, request.user.username)
                save_avto_history(car, "glonass_imei", old_imei, imei, request.user.username)
                save_avto_history(car, "glonass_iccid", old_iccid, iccid, request.user.username)

                car.glonass = glonass_text
                car.glonass_sn = sn
                car.glonass_imei = imei
                car.glonass_iccid = iccid

                if not old_glonass:
                    car.glonass_kto = request.user.username
                    car.glonass_kogda = timezone.now()
                    car.save(update_fields=[
                        "glonass",
                        "glonass_sn",
                        "glonass_imei",
                        "glonass_iccid",
                        "glonass_kto",
                        "glonass_kogda",
                    ])
                else:
                    car.save(update_fields=[
                        "glonass",
                        "glonass_sn",
                        "glonass_imei",
                        "glonass_iccid",
                    ])

                if old_glonass and is_manager:
                    messages.success(request, "Глонасс успешно изменен.")
                else:
                    messages.success(request, "Глонасс успешно привязан.")

                return redirect(f"/glonass/?car_id={car.id}")

    return render(request, "defects_app/glonass.html", {
        "form": form,
        "glonass_form": glonass_form,
        "car": car,
        "vin_prefixes": VIN_PREFIXES,
        "locked": locked,
        "is_manager": is_manager,
    })


# glonas i batareia vmeste

@login_required
def telematika_glonass_view(request):
    is_manager = is_manager_user(request.user)

    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]

    # Разрешаем доступ станции телематики, станции ГЛОНАСС и начальнику
    if not can_fix_aggregates(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    form = CarSearchForm()
    telematika_form = TelematikaForm()
    glonass_form = GlonassForm()

    car = None
    telematika_locked = False
    glonass_locked = False

    car_id = request.GET.get("car_id")

    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)

        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        if car.telematika:
            telematika_form = TelematikaForm(initial={
                "telematika": car.telematika
            })
            if not can_edit_aggregates(request.user):
                telematika_locked = True

        if car.glonass:
            glonass_form = GlonassForm(initial={
                "glonass": car.glonass
            })
            if not can_edit_aggregates(request.user):
                glonass_locked = True

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/telematika-glonass/?car_id={car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")

        elif action == "save_telematika":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            if car.telematika and not is_manager:
                messages.error(request, "Телематика уже привязана. Изменение доступно только начальнику.")
                return redirect(f"/telematika-glonass/?car_id={car.id}")

            telematika_form = TelematikaForm(request.POST)

            if telematika_form.is_valid():
                old_telematika = car.telematika
                new_telematika = telematika_form.cleaned_data["telematika"]

                save_avto_history(
                    car,
                    "telematika",
                    old_telematika,
                    new_telematika,
                    request.user.username
                )

                car.telematika = new_telematika

                if not old_telematika:
                    car.privyazal_telematiku = request.user.username
                    car.data_privyazki_telematiki = timezone.now()
                    car.save(update_fields=[
                        "telematika",
                        "privyazal_telematiku",
                        "data_privyazki_telematiki",
                    ])
                    messages.success(request, "Телематика успешно привязана.")
                else:
                    car.save(update_fields=["telematika"])
                    messages.success(request, "Телематика успешно изменена.")

                return redirect(f"/telematika-glonass/?car_id={car.id}")

        elif action == "save_glonass":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            if car.glonass and not is_manager:
                messages.error(request, "ГЛОНАСС уже привязан. Изменение доступно только начальнику.")
                return redirect(f"/telematika-glonass/?car_id={car.id}")

            glonass_form = GlonassForm(request.POST)

            if glonass_form.is_valid():
                old_glonass = car.glonass
                old_sn = car.glonass_sn
                old_imei = car.glonass_imei
                old_iccid = car.glonass_iccid

                glonass_text = glonass_form.cleaned_data["glonass"]
                sn, imei, iccid = parse_glonass_data(glonass_text)

                save_avto_history(car, "glonass", old_glonass, glonass_text, request.user.username)
                save_avto_history(car, "glonass_sn", old_sn, sn, request.user.username)
                save_avto_history(car, "glonass_imei", old_imei, imei, request.user.username)
                save_avto_history(car, "glonass_iccid", old_iccid, iccid, request.user.username)

                car.glonass = glonass_text
                car.glonass_sn = sn
                car.glonass_imei = imei
                car.glonass_iccid = iccid

                if not old_glonass:
                    car.glonass_kto = request.user.username
                    car.glonass_kogda = timezone.now()
                    car.save(update_fields=[
                        "glonass",
                        "glonass_sn",
                        "glonass_imei",
                        "glonass_iccid",
                        "glonass_kto",
                        "glonass_kogda",
                    ])
                    messages.success(request, "ГЛОНАСС успешно привязан.")
                else:
                    car.save(update_fields=[
                        "glonass",
                        "glonass_sn",
                        "glonass_imei",
                        "glonass_iccid",
                    ])
                    messages.success(request, "ГЛОНАСС успешно изменен.")

                return redirect(f"/telematika-glonass/?car_id={car.id}")

    return render(request, "defects_app/telematika_glonass.html", {
        "form": form,
        "telematika_form": telematika_form,
        "glonass_form": glonass_form,
        "car": car,
        "vin_prefixes": VIN_PREFIXES,
        "telematika_locked": telematika_locked,
        "glonass_locked": glonass_locked,
        "is_manager": is_manager,
    })



@login_required
def batareya_view(request):
    return redirect("/agregaty/")


@login_required
def perednij_dvigatel_view(request):
    return redirect("/agregaty/")


@login_required
def zadnij_dvigatel_view(request):
    return redirect("/agregaty/")

@login_required
def agregaty_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    is_manager = is_manager_user(request.user)
    station_id = session_data["station_id"]

    if not can_fix_aggregates(request.user):
        return HttpResponseForbidden("У вас нет прав на работу с агрегатами.")

    form = CarSearchForm()
    car = None

    batareya_form = BatareyaForm()
    perednij_form = PerednijDvigatelForm()
    zadnij_form = ZadnijDvigatelForm()

    car_id = request.GET.get("car_id")

    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)

        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        batareya_form = BatareyaForm(initial={
            "batareya": car.batareya
        })

        perednij_form = PerednijDvigatelForm(initial={
            "perednij_dvigatel": car.perednij_dvigatel
        })

        zadnij_form = ZadnijDvigatelForm(initial={
            "zadnij_dvigatel": car.zadnij_dvigatel
        })

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    return redirect(f"/agregaty/?car_id={car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")

        elif action == "bind":
            car_id_post = request.POST.get("car_id")
            component = request.POST.get("component")

            car = get_object_or_404(Avtomobili, id=car_id_post)

            if component == "batareya":
                if car.batareya and not can_edit_aggregates(request.user):
                    messages.error(request, "Батарея уже привязана. Изменение доступно только начальнику.")
                    return redirect(f"/agregaty/?car_id={car.id}")

                batareya_form = BatareyaForm(request.POST)

                if batareya_form.is_valid():
                    old_value = car.batareya
                    new_value = batareya_form.cleaned_data["batareya"]

                    save_avto_history(car, "batareya", old_value, new_value, request.user.username)

                    car.batareya = new_value

                    if not old_value:
                        car.batareya_kto = request.user.username
                        car.batareya_kogda = timezone.now()
                        car.save(update_fields=["batareya", "batareya_kto", "batareya_kogda"])
                    else:
                        car.save(update_fields=["batareya"])

                    messages.success(request, "Батарея сохранена.")
                    return redirect(f"/agregaty/?car_id={car.id}")

            elif component == "perednij_dvigatel":
                if car.perednij_dvigatel and not can_edit_aggregates(request.user):
                    messages.error(request, "Передний двигатель уже привязан. Изменение доступно только начальнику.")
                    return redirect(f"/agregaty/?car_id={car.id}")

                perednij_form = PerednijDvigatelForm(request.POST)

                if perednij_form.is_valid():
                    old_value = car.perednij_dvigatel
                    new_value = perednij_form.cleaned_data["perednij_dvigatel"]

                    save_avto_history(car, "perednij_dvigatel", old_value, new_value, request.user.username)

                    car.perednij_dvigatel = new_value

                    if not old_value:
                        car.perednij_dvigatel_kto = request.user.username
                        car.perednij_dvigatel_kogda = timezone.now()
                        car.save(update_fields=[
                            "perednij_dvigatel",
                            "perednij_dvigatel_kto",
                            "perednij_dvigatel_kogda",
                        ])
                    else:
                        car.save(update_fields=["perednij_dvigatel"])

                    messages.success(request, "Передний двигатель сохранён.")
                    return redirect(f"/agregaty/?car_id={car.id}")

            elif component == "zadnij_dvigatel":
                if car.zadnij_dvigatel and not can_edit_aggregates(request.user):
                    messages.error(request, "Задний двигатель уже привязан. Изменение доступно только начальнику.")
                    return redirect(f"/agregaty/?car_id={car.id}")

                zadnij_form = ZadnijDvigatelForm(request.POST)

                if zadnij_form.is_valid():
                    old_value = car.zadnij_dvigatel
                    new_value = zadnij_form.cleaned_data["zadnij_dvigatel"]

                    save_avto_history(car, "zadnij_dvigatel", old_value, new_value, request.user.username)

                    car.zadnij_dvigatel = new_value

                    if not old_value:
                        car.zadnij_dvigatel_kto = request.user.username
                        car.zadnij_dvigatel_kogda = timezone.now()
                        car.save(update_fields=[
                            "zadnij_dvigatel",
                            "zadnij_dvigatel_kto",
                            "zadnij_dvigatel_kogda",
                        ])
                    else:
                        car.save(update_fields=["zadnij_dvigatel"])

                    messages.success(request, "Задний двигатель сохранён.")
                    return redirect(f"/agregaty/?car_id={car.id}")

    return render(request, "defects_app/agregaty.html", {
        "form": form,
        "car": car,
        "batareya_form": batareya_form,
        "perednij_form": perednij_form,
        "zadnij_form": zadnij_form,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager,
    })

@login_required
def quality_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]
    is_manager = can_edit_delete_defects(request.user)

    if not can_create_defects(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    form = CarSearchForm()
    car = None
    defects = []

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })
        defects = Defekty.objects.filter(avto=car).order_by("-data")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                existing_car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if existing_car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/quality/?car_id={existing_car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")
                    return render(request, "defects_app/quality.html", {
                        "form": form,
                        "car": None,
                        "defects": [],
                        "vin_prefixes": VIN_PREFIXES,
                        "is_manager": is_manager,
                    })

    return render(request, "defects_app/quality.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager,
        "from_page": "quality",
    })


@login_required
def vh1_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    station_id = session_data["station_id"]
    is_manager = can_edit_delete_defects(request.user)

    station_name = session_data.get("station_name")
    if not is_vh1_station(station_id=station_id, station_name=station_name):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    if not can_create_defects(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    form = CarSearchForm()
    car = None
    defects = []

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })
        defects = Defekty.objects.filter(avto=car).order_by("-data")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                existing_car = Avtomobili.objects.filter(vin=vin).first()

                if existing_car:
                    messages.success(request, "Машина найдена.")
                    return redirect(f"/vh1/?car_id={existing_car.id}")

                plan_vin = PlanovyeVin.objects.filter(vin=vin).first()
                if not plan_vin:
                    messages.error(request, "Машина с таким VIN не найдена в созданных и плановых VIN.")
                    return render(request, "defects_app/quality.html", {
                        "form": form,
                        "car": None,
                        "defects": [],
                        "vin_prefixes": VIN_PREFIXES,
                        "is_manager": is_manager,
                        "from_page": "vh1",
                    })

                model_name = plan_vin.model.strip()
                model_obj, _ = Modeli.objects.get_or_create(nazvanie=model_name)

                created_car = Avtomobili.objects.create(
                    vin=vin,
                    model=model_obj,
                    kto_sozdal=request.user.username,
                    data_sozdaniya=timezone.now(),
                    created_on_station_1=False,
                )
                messages.success(request, "Машина найдена в плановых VIN и создана для ВХ1.")
                return redirect(f"/vh1/?car_id={created_car.id}")

    return render(request, "defects_app/quality.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager,
        "from_page": "vh1",
    })


@login_required
def dovodka_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response
    
    is_manager = is_manager_user(request.user)

    station_id = session_data["station_id"]
    if not can_create_defects(request.user):
        return HttpResponseForbidden("Для вашей станции эта форма недоступна.")

    form = CarSearchForm()
    car = None
    defects = []
    snp_defects = []
    all_verified = False

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })
        defects = Defekty.objects.filter(avto=car).prefetch_related(
            "snp_comments"
        ).order_by("-data")

        snp_defects = Defekty.objects.filter(
            avto=car,
            ustraneno=False
        ).order_by("-data")

        all_verified = defects.exists() and all(defect.proveren for defect in defects)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)
            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if car:
                    return redirect(f"/dovodka/?car_id={car.id}")
                else:
                    messages.error(request, "Машина с таким VIN не найдена.")

        elif action == "save_repairs":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)

            defects = Defekty.objects.filter(avto=car).order_by("-data")
            fixed_ids = request.POST.getlist("fixed_defects")

            fixed_count = 0

            for defect in defects:
                if not defect.ustraneno and str(defect.id) in fixed_ids:
                    defect.ustraneno = True
                    defect.kto_ustranil = request.user.username
                    defect.data_ustraneniya = timezone.now()
                    defect.save()
                    fixed_count += 1

            if fixed_count > 0:
                messages.success(request, f"Отмечено устраненных дефектов: {fixed_count}.")
            else:
                messages.info(request, "Новых отметок устранения не было.")

            return redirect(f"/dovodka/?car_id={car.id}")
        
        elif action == "send_snp":
            car_id_post = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id_post)
            defects = Defekty.objects.filter(avto=car)

            if not car_passed_bestenevaya(car):
                messages.error(
                    request,
                    "Нельзя отправить машину на СНП. Сначала машина должна пройти Бестеневую."
                )
                return redirect(f"/dovodka/?car_id={car.id}")

            last_status = StatusAvto.objects.filter(avto=car).order_by("-data_statusa").first()

            if last_status and last_status.status == "СГП":
                messages.error(request, "Машина уже находится в СГП. На СНП её отправить нельзя.")
                return redirect(f"/dovodka/?car_id={car.id}")

            all_verified = defects.exists() and all(defect.proveren for defect in defects)
            if all_verified:
                messages.error(request, "Машина уже полностью проверена. На СНП её отправлять нельзя.")
                return redirect(f"/dovodka/?car_id={car.id}")

            snp_defects = Defekty.objects.filter(avto=car, ustraneno=False).order_by("-data")

            if snp_defects.exists():
                has_empty_comment = False

                for defect in snp_defects:
                    comment_text = request.POST.get(f"snp_comment_{defect.id}", "").strip()

                    if not comment_text:
                        has_empty_comment = True
                        break

                if has_empty_comment:
                    messages.error(request, "Для каждого неустраненного дефекта нужно указать причину отправки на СНП.")
                    return redirect(f"/dovodka/?car_id={car.id}")

                for defect in snp_defects:
                    comment_text = request.POST.get(f"snp_comment_{defect.id}", "").strip()

                    SnpDefectComment.objects.create(
                        defect=defect,
                        avto=car,
                        comment=comment_text,
                        kto_sozdal=request.user.username,
                    )

            StatusAvto.objects.create(
                avto=car,
                status="СНП",
                kto_izmenil=request.user.username
            )

            messages.success(request, "Машина передана на СНП.")

            return redirect(f"/dovodka/?car_id={car.id}")

    return render(request, "defects_app/dovodka.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager_user(request.user),
        "snp_defects": snp_defects,
        "all_verified": all_verified
    })


@login_required
def create_car_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    is_manager = is_manager_user(request.user)

    station_id = session_data["station_id"]
    if not can_create_cars_and_print(request.user):
        return HttpResponseForbidden("У вас нет прав на создание машины и печать.")

    form = CarSearchForm()
    search_form = CarSearchForm()

    car = None
    search_result_car = None
    search_result_plan = None
    search_not_found = False
    show_search_modal = False

    last_created_cars_raw = Avtomobili.objects.order_by("-data_sozdaniya")[:10]

    important_sheet_count = len(request.session.get("important_sheet_car_ids", []))

    last_created_cars = []
    for created_car in last_created_cars_raw:
        plan = PlanovyeVin.objects.filter(vin=created_car.vin).first()
        last_created_cars.append({
            "car": created_car,
            "plan": plan,
        })

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "search_created_car":
            search_form = CarSearchForm(request.POST)
            show_search_modal = True

            if search_form.is_valid():
                vin = search_form.cleaned_data["vin"]

                search_result_car = Avtomobili.objects.filter(vin=vin).first()

                if search_result_car:
                    search_result_plan = PlanovyeVin.objects.filter(vin=search_result_car.vin).first()
                else:
                    search_not_found = True

            return render(request, "defects_app/create_car.html", {
                "form": form,
                "search_form": search_form,
                "car": car,
                "vin_prefixes": VIN_PREFIXES,
                "is_manager": is_manager,
                "last_created_cars": last_created_cars,
                "search_result_car": search_result_car,
                "search_result_plan": search_result_plan,
                "search_not_found": search_not_found,
                "show_search_modal": show_search_modal,
                "important_sheet_count": important_sheet_count,
            })

        form = CarSearchForm(request.POST)

        if form.is_valid():
            vin = form.cleaned_data["vin"]
            model = form.cleaned_data["model"]

            plan_vin = PlanovyeVin.objects.filter(vin=vin).first()

            if not plan_vin:
                messages.error(
                    request,
                    "Такой VIN-номер не был предварительно внесен в базу. Обратитесь к ответственному за формирование VIN-номеров."
                )
                return render(request, "defects_app/create_car.html", {
                    "form": form,
                    "search_form": search_form,
                    "car": None,
                    "vin_prefixes": VIN_PREFIXES,
                    "is_manager": is_manager,
                    "last_created_cars": last_created_cars,
                    "search_result_car": search_result_car,
                    "search_result_plan": search_result_plan,
                    "search_not_found": search_not_found,
                    "show_search_modal": show_search_modal,
                    "important_sheet_count": important_sheet_count,
                })

            existing_car = Avtomobili.objects.filter(vin=vin).first()

            if existing_car:
                if not existing_car.created_on_station_1:
                    existing_car.created_on_station_1 = True
                    existing_car.kto_sozdal = request.user.username
                    existing_car.data_sozdaniya = timezone.now()
                    existing_car.save(update_fields=[
                        "created_on_station_1",
                        "kto_sozdal",
                        "data_sozdaniya",
                    ])

                    messages.success(request, "Машина успешно создана на станции 1.")
                    return redirect("print_created_car", car_id=existing_car.id)

                messages.info(request, "Машина с таким VIN уже существует.")
                return redirect(f"/create-car/?car_id={existing_car.id}")

            model_name = plan_vin.model.strip()

            model_obj, created = Modeli.objects.get_or_create(
                nazvanie=model_name
            )

            car = Avtomobili.objects.create(
                vin=vin,
                model=model_obj,
                kto_sozdal=request.user.username,
                data_sozdaniya=timezone.now(),
            )

            messages.success(request, "Новая машина успешно создана.")
            return redirect("print_created_car", car_id=car.id)

    car_id = request.GET.get("car_id")
    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)
        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

    return render(request, "defects_app/create_car.html", {
        "form": form,
        "search_form": search_form,
        "car": car,
        "vin_prefixes": VIN_PREFIXES,
        "is_manager": is_manager,
        "last_created_cars": last_created_cars,
        "search_result_car": search_result_car,
        "search_result_plan": search_result_plan,
        "search_not_found": search_not_found,
        "show_search_modal": show_search_modal,
        "important_sheet_count": important_sheet_count,
    })


@login_required
def manager_dashboard_view(request):
    response = manager_required_view(request)
    if response:
        return response

    shift_name = request.session.get("shift_name", "")

    return render(request, "defects_app/manager_dashboard.html", {
        "shift_name": shift_name,
    })


def manager_login_view(request):
    if request.user.is_authenticated:
        if is_manager_user(request.user):
            return redirect("manager_dashboard")
        logout(request)

    if request.method == "POST":
        form = ManagerLoginForm(request, data=request.POST)

        if form.is_valid():
            user = form.get_user()

            if not is_manager_user(user):
                messages.error(request, "Этот вход доступен только начальникам.")
                return render(request, "registration/manager_login.html", {"form": form})

            shift = form.cleaned_data["shift"]

            login(request, user)

            request.session["station_id"] = 999
            request.session["station_name"] = "Панель начальника"
            request.session["shift_id"] = shift.id
            request.session["shift_name"] = shift.nazvanie

            return redirect("manager_dashboard")
        else:
            messages.error(request, "Проверьте логин, пароль и смену.")
    else:
        form = ManagerLoginForm(request)

    return render(request, "registration/manager_login.html", {"form": form})

def generate_vin_barcode_svg(vin):
    code128 = barcode.get("code128", vin, writer=SVGWriter())
    buffer = BytesIO()
    code128.write(buffer, options={
        "write_text": True,
        "module_height": 18,
        "font_size": 10,
        "quiet_zone": 2,
    })
    return buffer.getvalue().decode("utf-8")


@login_required
def print_created_car_view(request, car_id):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response
    if not can_create_cars_and_print(request.user):
        return HttpResponseForbidden("У вас нет прав на печать карточек автомобиля.")

    car = get_object_or_404(Avtomobili, id=car_id)

    plan_vin = PlanovyeVin.objects.filter(vin=car.vin).first()
    vin_barcode_svg = generate_vin_barcode_svg(car.vin)
    kolichestvo_mest = "7" if "C2A7" in car.vin or (plan_vin and "/7_" in plan_vin.komplektaciya) else "5"

    stations = [ ]
    #     "1. Бестеневая",
    #     "2. OkLine",
    #     "3. Телематика",
    #     "4. Глонасс",
    #     "5. Батарея",
    #     "6. Передний двигатель",
    #     "7. Задний двигатель",
    #     "8. Станция 3 – Качество",
    #     "9. T20 – Качество",
    #     "10. C06 – Качество",
    #     "11. Доводка",
    # ]

    aggregates = [
        "Телематика",
        "Глонасс",
        "Батарея",
        "Передний двигатель",
        "Задний двигатель",
    ]

    important_print_url = None
    if session_data.get("station_id") in [1, 12, 999]:
        buffer_key = "important_sheet_car_ids"
        car_ids_buffer = request.session.get(buffer_key, [])
        if car.id not in car_ids_buffer:
            car_ids_buffer.append(car.id)

        if len(car_ids_buffer) >= 6:
            request.session[buffer_key] = []
            important_print_url = reverse("print_created_car_info_batch") + "?" + urlencode({
                "car_ids": ",".join(str(buffer_car_id) for buffer_car_id in car_ids_buffer[:6])
            })
        else:
            request.session[buffer_key] = car_ids_buffer

        request.session.modified = True

    return render(request, "defects_app/print_created_car.html", {
        "car": car,
        "stations": stations,
        "aggregates": aggregates,
        "plan_vin": plan_vin,
        "vin_barcode_svg": vin_barcode_svg,
        "kolichestvo_mest": kolichestvo_mest,
        "important_print_url": important_print_url,
    })

@login_required
def print_created_car_info_view(request, car_id):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response
    if not can_create_cars_and_print(request.user):
        return HttpResponseForbidden("У вас нет прав на печать карточек автомобиля.")

    car = get_object_or_404(Avtomobili, id=car_id)
    plan_vin = PlanovyeVin.objects.filter(vin=car.vin).first()
    vin_barcode_svg = generate_vin_barcode_svg(car.vin)
    kolichestvo_mest = "7" if "C2A7" in car.vin or (plan_vin and "/7_" in plan_vin.komplektaciya) else "5"

    return render(request, "defects_app/print_created_car_info.html", {
        "car": car,
        "plan_vin": plan_vin,
        "vin_barcode_svg": vin_barcode_svg,
        "kolichestvo_mest": kolichestvo_mest,
    })


@login_required
def print_created_car_info_batch_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response
    if not can_create_cars_and_print(request.user):
        return HttpResponseForbidden("У вас нет прав на печать карточек автомобиля.")

    car_ids_raw = request.GET.get("car_ids", "")
    car_ids = [int(car_id.strip()) for car_id in car_ids_raw.split(",") if car_id.strip().isdigit()][:6]

    if not car_ids:
        return HttpResponseBadRequest("Не переданы данные для печати важного листа.")

    cars = Avtomobili.objects.filter(id__in=car_ids)
    cars_by_id = {car.id: car for car in cars}

    print_rows = []
    for car_id in car_ids:
        car = cars_by_id.get(car_id)
        if not car:
            continue
        plan_vin = PlanovyeVin.objects.filter(vin=car.vin).first()
        kolichestvo_mest = "7" if "C2A7" in car.vin or (plan_vin and "/7_" in plan_vin.komplektaciya) else "5"
        print_rows.append({
            "car": car,
            "plan_vin": plan_vin,
            "vin_barcode_svg": generate_vin_barcode_svg(car.vin),
            "kolichestvo_mest": kolichestvo_mest,
        })

    if not print_rows:
        return HttpResponseBadRequest("Не удалось подготовить данные для печати важного листа.")

    return render(request, "defects_app/print_created_car_info_batch.html", {
        "print_rows": print_rows,
    })

@login_required
def print_sgp_report_view(request, car_id):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    car = get_object_or_404(Avtomobili, id=car_id)

    defects = Defekty.objects.filter(avto=car).select_related(
        "mesto",
        "tip",
        "greyd",
        "otvetstvennyj",
        "sgp_approval",
    ).order_by("data")

    aggregates = [
        {"name": "Телематика", "value": car.telematika or ""},
        {"name": "Глонасс", "value": car.glonass or ""},
        {"name": "Батарея", "value": car.batareya or ""},
        {"name": "Передний двигатель", "value": car.perednij_dvigatel or ""},
        {"name": "Задний двигатель", "value": car.zadnij_dvigatel or ""},
    ]

    return render(request, "defects_app/print_sgp_report.html", {
        "car": car,
        "defects": defects,
        "aggregates": aggregates,
    })


@login_required
def manager_open_station_view(request, station_id):
    response = manager_required_view(request)
    if response:
        return response

    station = Mesta.objects.filter(id=station_id).first()
    if not station:
        if station_id == 13:
            vh1_station = get_vh1_station()
            if vh1_station:
                request.session["station_id"] = vh1_station.id
                request.session["station_name"] = vh1_station.nazvanie
                return redirect("vh1")
        raise Http404("No Mesta matches the given query.")

    request.session["station_id"] = station.id
    request.session["station_name"] = station.nazvanie

    redirect_name = get_station_redirect_name(station.id)
    return redirect(redirect_name)


@login_required
def manager_open_vh1_station_view(request):
    response = manager_required_view(request)
    if response:
        return response

    station = get_vh1_station()
    if not station:
        messages.error(request, "Станция ВХ1 не найдена в справочнике мест.")
        return redirect("manager_dashboard")

    request.session["station_id"] = station.id
    request.session["station_name"] = station.nazvanie
    return redirect("vh1")

@login_required
def open_station_for_department_view(request, station_id):
    station = Mesta.objects.filter(id=station_id).first()
    if not station:
        if station_id == 13:
            return redirect("open_vh1_station_for_department")
        raise Http404("No Mesta matches the given query.")

    user = request.user
    allowed = False

    # Глобальные начальники могут всё
    if is_manager_user(user):
        allowed = True
    else:
        # ОТК
        if station_id in [1, 2, 8, 9, 10, 11] and can_create_defects(user):
            allowed = True

        if can_create_defects(user) and is_vh1_station(station_id=station.id, station_name=station.nazvanie):
            allowed = True

        # Агрегаты
        if station_id in [3, 4, 5, 6, 7] and can_fix_aggregates(user):
            allowed = True

        # Логистика (станция создания машины)
        if station_id == 12 and can_create_cars_and_print(user):
            allowed = True

    if not allowed:
        return render_access_denied(request, "У вас нет доступа к выбранной станции для этого отдела.")

    # Смена уже должна быть выбрана на логине
    shift_id = request.session.get("shift_id")
    shift_name = request.session.get("shift_name")
    if not shift_id or not shift_name:
        logout(request)
        messages.error(request, "Смена не выбрана. Войдите заново.")
        return redirect("login")

    request.session["station_id"] = station.id
    request.session["station_name"] = station.nazvanie

    redirect_name = get_station_redirect_name(station.id)
    return redirect(redirect_name)


@login_required
def open_vh1_station_for_department_view(request):
    station = get_vh1_station()
    if not station:
        return render_access_denied(request, "Станция ВХ1 не найдена в справочнике мест.")

    user = request.user
    allowed = False
    if is_manager_user(user):
        allowed = True
    elif can_create_defects(user):
        allowed = True

    if not allowed:
        return render_access_denied(request, "У вас нет доступа к выбранной станции для этого отдела.")

    shift_id = request.session.get("shift_id")
    shift_name = request.session.get("shift_name")
    if not shift_id or not shift_name:
        logout(request)
        messages.error(request, "Смена не выбрана. Войдите заново.")
        return redirect("login")

    request.session["station_id"] = station.id
    request.session["station_name"] = station.nazvanie
    return redirect("vh1")

@login_required
def snp_orders_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    is_manager = is_manager_user(request.user)

    form = CarSearchForm()
    car = None
    defects = []
    message_text = None

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "find":
            form = CarSearchForm(request.POST)

            if form.is_valid():
                vin = form.cleaned_data["vin"]
                car = Avtomobili.objects.filter(
                    vin=vin,
                    created_on_station_1=True
                ).first()

                if not car:
                    messages.error(request, "Машина с таким VIN не найдена.")
                    return redirect("snp_orders")

                last_status = StatusAvto.objects.filter(avto=car).order_by("-data_statusa").first()

                if not last_status or last_status.status != "СНП":
                    messages.error(request, "Эта машина сейчас не находится на СНП. Ввод номеров заказов недоступен.")
                    return redirect("snp_orders")

                return redirect(f"/snp-orders/?car_id={car.id}")

        elif action == "save_orders":
            car_id = request.POST.get("car_id")
            car = get_object_or_404(Avtomobili, id=car_id)

            last_status = StatusAvto.objects.filter(avto=car).order_by("-data_statusa").first()

            if not last_status or last_status.status != "СНП":
                messages.error(request, "Сохранять заказы можно только для машины, которая сейчас находится на СНП.")
                return redirect("snp_orders")

            defects = Defekty.objects.filter(avto=car).prefetch_related("snp_order")

            saved_count = 0
            updated_count = 0

            for defect in defects:
                if defect.ustraneno:
                    continue

                order_number = request.POST.get(f"order_{defect.id}", "").strip()

                if not order_number:
                    continue

                existing_order = getattr(defect, "snp_order", None)

                if existing_order:
                    if not can_edit_delete_defects(request.user):
                        continue

                    if existing_order.order_number != order_number:
                        existing_order.order_number = order_number
                        existing_order.kto_izmenil = request.user.username
                        existing_order.data_izmeneniya = timezone.now()
                        existing_order.save()
                        updated_count += 1
                else:
                    SnpDefectOrder.objects.create(
                        defect=defect,
                        avto=car,
                        order_number=order_number,
                        kto_sozdal=request.user.username,
                    )
                    saved_count += 1

            messages.success(request, f"Сохранено новых заказов: {saved_count}. Изменено заказов: {updated_count}.")
            return redirect(f"/snp-orders/?car_id={car.id}")

    car_id = request.GET.get("car_id")

    if car_id:
        car = get_object_or_404(Avtomobili, id=car_id)

        last_status = StatusAvto.objects.filter(avto=car).order_by("-data_statusa").first()

        if not last_status or last_status.status != "СНП":
            messages.error(request, "Эта машина сейчас не находится на СНП. Ввод номеров заказов недоступен.")
            return redirect("snp_orders")

        form = CarSearchForm(initial={
            "vin": car.vin,
            "model": car.model
        })

        defects = Defekty.objects.filter(avto=car).select_related(
            "tip",
            "oblast",
            "greyd",
            "otvetstvennyj",
            "mesto"
        ).prefetch_related(
            "snp_order"
        ).order_by("id")

    return render(request, "defects_app/snp_orders.html", {
        "form": form,
        "car": car,
        "defects": defects,
        "is_manager": is_manager,
        "vin_prefixes": VIN_PREFIXES,
    })


@login_required
def upload_plan_vin_view(request):
    if not (is_manager_user(request.user) or is_planning_user(request.user)):
        return HttpResponseForbidden("Доступ только для начальников и отдела планирования производства.")

    preview_data = None

    last_record = PlanovyeVin.objects.exclude(otts="").order_by("-data_zagruzki").first()
    last_otts = last_record.otts if last_record else ""

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "check":
            excel_file = request.FILES.get("excel_file")
            otts = request.POST.get("otts", "").strip()

            if not excel_file:
                messages.error(request, "Выберите Excel-файл.")
                return redirect("upload_plan_vin")

            if not otts:
                messages.error(request, "Заполните ОТТС.")
                return redirect("upload_plan_vin")

            preview_data = check_plan_vin_excel(excel_file)
            preview_data["otts"] = otts

            request.session["plan_vin_preview"] = preview_data

        elif action == "import":
            preview_data = request.session.get("plan_vin_preview")

            if not preview_data:
                messages.error(request, "Сначала проверьте файл.")
                return redirect("upload_plan_vin")

            if preview_data["errors"]:
                messages.error(request, "Файл содержит ошибки. Загрузка отменена.")
                return redirect("upload_plan_vin")

            otts = preview_data.get("otts", "").strip()

            if not otts:
                messages.error(request, "ОТТС не найден. Проверьте файл заново.")
                return redirect("upload_plan_vin")

            created_count = 0
            updated_count = 0

            for item in preview_data["rows"]:
                obj, created = PlanovyeVin.objects.update_or_create(
                    vin=item["vin"],
                    defaults={
                        "nomer_partii": item["nomer_partii"],
                        "nomer_lota": item["nomer_lota"],
                        "model": item["model"],
                        "cvet_kuzova": item["cvet_kuzova"],
                        "cvet_salona": item["cvet_salona"],
                        "komplektaciya": item["komplektaciya"],
                        "otts": otts,
                        "file_name": preview_data["file_name"],
                        "kto_zagruzil": request.user.username,
                        "data_zagruzki": timezone.now(),
                        "otts": preview_data.get("otts"),
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            request.session.pop("plan_vin_preview", None)

            messages.success(
                request,
                f"Файл загружен. Новых VIN: {created_count}. Обновлено: {updated_count}."
            )
            return redirect("upload_plan_vin")

        elif action == "cancel":
            request.session.pop("plan_vin_preview", None)
            return redirect("upload_plan_vin")

    total_count = PlanovyeVin.objects.count()
    last_items = PlanovyeVin.objects.order_by("-data_zagruzki")[:20]

    return render(request, "defects_app/upload_plan_vin.html", {
        "total_count": total_count,
        "last_items": last_items,
        "preview_data": preview_data,
        "last_otts": last_otts,
    })


def check_plan_vin_excel(excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    header_row = None

    for row in ws.iter_rows(min_row=1, max_row=20):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        if "Вин рус" in values:
            header_row = row[0].row
            break

    if not header_row:
        return {
            "file_name": excel_file.name,
            "total_rows": 0,
            "rows": [],
            "errors": ["Не найдена строка заголовков. В файле должен быть столбец «Вин рус»."],
            "warnings": [],
        }

    rows_for_import = []
    errors = []
    warnings = []
    vins_in_file = set()

    total_rows = 0

    for excel_row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1
    ):
        if not row or all(value is None for value in row):
            continue

        total_rows += 1

        nomer_partii = row[1]
        nomer_lota = row[2]
        model = row[3]
        vin = row[4]
        cvet_kuzova = row[5]
        cvet_salona = row[6]
        komplektaciya = row[7]

        vin = str(vin).strip().upper() if vin else ""
        komplektaciya_text = str(komplektaciya).strip() if komplektaciya else ""

        if not vin:
            errors.append(f"Строка {excel_row_number}: VIN пустой.")
            continue

        if len(vin) != 17:
            errors.append(f"Строка {excel_row_number}: VIN «{vin}» должен содержать 17 символов.")
            continue

        if vin in vins_in_file:
            errors.append(f"Строка {excel_row_number}: VIN «{vin}» повторяется внутри файла.")
            continue

        vins_in_file.add(vin)

        if "/7_" in komplektaciya_text and "C2A7" not in vin:
            errors.append(
                f"Строка {excel_row_number}: комплектация «{komplektaciya_text}», но VIN «{vin}» не семиместный."
            )

        if "/5_" in komplektaciya_text and ("C2A5" not in vin and "C2B5" not in vin):
            errors.append(
                f"Строка {excel_row_number}: комплектация «{komplektaciya_text}», но VIN «{vin}» не пятиместный."
            )

        if PlanovyeVin.objects.filter(vin=vin).exists():
            warnings.append(f"Строка {excel_row_number}: VIN «{vin}» уже есть в базе и будет обновлен.")

        rows_for_import.append({
            "nomer_partii": str(nomer_partii).strip() if nomer_partii is not None else "",
            "nomer_lota": str(nomer_lota).strip() if nomer_lota is not None else "",
            "model": str(model).strip() if model is not None else "",
            "vin": vin,
            "cvet_kuzova": str(cvet_kuzova).strip() if cvet_kuzova is not None else "",
            "cvet_salona": str(cvet_salona).strip() if cvet_salona is not None else "",
            "komplektaciya": komplektaciya_text,
        })

    return {
        "file_name": excel_file.name,
        "total_rows": total_rows,
        "rows": rows_for_import,
        "errors": errors,
        "warnings": warnings,
    }

@login_required
def live_cars_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return redirect_response

    cars_raw = Avtomobili.objects.filter(
        proshla_bestenevaya=False
    ).order_by("data_sozdaniya")

    cars = []

    for car in cars_raw:
        plan = PlanovyeVin.objects.filter(vin=car.vin).first()

        kolichestvo_mest = "5"

        if plan:
            komplektaciya = plan.komplektaciya or ""
            if "C2A7" in car.vin or "/7" in komplektaciya:
                kolichestvo_mest = "7"

        cars.append({
            "car": car,
            "plan": plan,
            "kolichestvo_mest": kolichestvo_mest,
        })

    return render(request, "defects_app/live_cars.html", {
        "cars": cars,
    })

@login_required
def logistics_live_cars_view(request):
    if not (is_log_worker(request.user) or is_log_manager(request.user) or is_manager_user(request.user)):
        return render_access_denied(request, "У вас нет доступа к экрану логистики.")

    shift_id = request.session.get("shift_id")
    shift_name = request.session.get("shift_name")
    if not shift_id or not shift_name:
        logout(request)
        messages.info(request, "Сессия истекла. Войдите снова.")
        return redirect("login")

    cars_raw = Avtomobili.objects.filter(
        proshla_bestenevaya=False
    ).order_by("data_sozdaniya")

    cars = []

    for car in cars_raw:
        plan = PlanovyeVin.objects.filter(vin=car.vin).first()

        kolichestvo_mest = "5"

        if plan:
            komplektaciya = plan.komplektaciya or ""
            if "C2A7" in car.vin or "/7" in komplektaciya:
                kolichestvo_mest = "7"

        cars.append({
            "car": car,
            "plan": plan,
            "kolichestvo_mest": kolichestvo_mest,
        })

    return render(request, "defects_app/logistics_live_cars.html", {
        "cars": cars,
    })

def get_export_period(request, redirect_name):
    period = request.POST.get("period")
    now = timezone.now()

    if period == "today":
        start_datetime = datetime.combine(now.date(), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "yesterday":
        yesterday = now.date() - timedelta(days=1)
        start_datetime = datetime.combine(yesterday, time.min)
        end_datetime = datetime.combine(yesterday, time.max)

    elif period == "shift_today":
        start_datetime = datetime.combine(now.date(), time(8, 0))
        end_datetime = datetime.combine(now.date(), time(17, 0))

    elif period == "shift_yesterday":
        yesterday = now.date() - timedelta(days=1)
        start_datetime = datetime.combine(yesterday, time(8, 0))
        end_datetime = datetime.combine(yesterday, time(17, 0))

    elif period == "week":
        start_datetime = datetime.combine(now.date() - timedelta(days=7), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "month":
        start_datetime = datetime.combine(now.date() - timedelta(days=30), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "custom":
        date_from = request.POST.get("date_from")
        time_from = request.POST.get("time_from") or "00:00"
        date_to = request.POST.get("date_to")
        time_to = request.POST.get("time_to") or "23:59"

        if not date_from or not date_to:
            messages.error(request, "Для своей выгрузки нужно выбрать дату начала и дату окончания.")
            return None, None, redirect(redirect_name)

        start_datetime = datetime.strptime(f"{date_from} {time_from}", "%Y-%m-%d %H:%M")
        end_datetime = datetime.strptime(f"{date_to} {time_to}", "%Y-%m-%d %H:%M")

    else:
        messages.error(request, "Выберите период выгрузки.")
        return None, None, redirect(redirect_name)

    return start_datetime, end_datetime, None


def format_datetime_for_excel(value):
    return value.strftime("%d.%m.%Y %H:%M:%S") if value else ""


def format_duration_for_excel(start, end):
    if not start or not end:
        return ""

    diff = end - start
    total_seconds = int(diff.total_seconds())

    if total_seconds < 0:
        return ""

    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    if days > 0:
        return f"{days} д. {hours:02}:{minutes:02}:{seconds:02}"

    return f"{hours:02}:{minutes:02}:{seconds:02}"


def apply_excel_style(ws, column_widths):
    header_fill = PatternFill("solid", fgColor="B9A7CC")
    header_font = Font(name="Bahnschrift SemiLight", size=12, bold=True, color="000000")
    body_font = Font(name="Bahnschrift SemiLight", size=11, color="000000")

    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=False
            )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center",
            wrap_text=False
        )

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 22


@login_required
def exports_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")    

    response = manager_required_view(request)
    if response:
        return response

    return render(request, "defects_app/exports.html")


@login_required
def export_defects_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")

    response = manager_required_view(request)
    if response:
        return response

    if request.method == "POST":
        start_datetime, end_datetime, redirect_response = get_export_period(request, "export_defects")
        if redirect_response:
            return redirect_response

        defects = Defekty.objects.filter(
            data__gte=start_datetime,
            data__lte=end_datetime
        ).select_related(
            "avto",
            "avto__model",
            "tip",
            "oblast",
            "greyd",
            "otvetstvennyj",
            "smena",
            "mesto",
        ).prefetch_related(
            "photos"
        ).order_by("data", "avto__model__nazvanie")

        wb = Workbook()
        ws = wb.active
        ws.title = "Дефекты"

        headers = [
            "Дата внесения дефекта",
            "ID дефекта",
            "Станция",
            "VIN",
            "Модель",
            "Тип дефекта",
            "Область дефекта",
            "Грейд",
            "Ответственный",
            "Комментарий",
            "Устранена",
            "Кто устранил",
            "Дата устранения",
            "Проверена",
            "Кто разрешил",
            "Дата разрешения",
            "Дата отправки на СГП",
            "Смена",
            "Фото дефекта",
        ]

        ws.append(headers)

        for defect in defects:
            sgp_status = StatusAvto.objects.filter(
                avto=defect.avto,
                status="СГП"
            ).order_by("-data_statusa").first()

            photo_count = defect.photos.count()

            photo_count = defect.photos.count()

            if photo_count > 0:
                defect_url = f"http://defects-auto.irito.ru/defect-photo-detail/{defect.id}/"
                photo_cell_text = "Посмотреть фото"
            else:
                defect_url = None
                photo_cell_text = ""

            ws.append([
                format_datetime_for_excel(defect.data),
                defect.id,
                defect.mesto.nazvanie if defect.mesto else "",
                defect.avto.vin if defect.avto else "",
                defect.avto.model.nazvanie if defect.avto and defect.avto.model else "",
                defect.tip.nazvanie if defect.tip else "",
                defect.oblast.nazvanie if defect.oblast else "",
                defect.greyd.nazvanie if defect.greyd else "",
                defect.otvetstvennyj.nazvanie if defect.otvetstvennyj else "",
                defect.kommentarij or "",
                "Да" if defect.ustraneno else "Нет",
                defect.kto_ustranil or "",
                format_datetime_for_excel(defect.data_ustraneniya),
                "Да" if defect.proveren else "Нет",
                defect.kto_razreshil or "",
                format_datetime_for_excel(defect.data_razresheniya),
                format_datetime_for_excel(sgp_status.data_statusa) if sgp_status else "",
                defect.smena.nazvanie if defect.smena else "",
                photo_cell_text,
            ])

            current_row = ws.max_row

            if defect_url:
                cell = ws[f"S{current_row}"]
                cell.hyperlink = defect_url
                cell.style = "Hyperlink"


        column_widths = {
            "A": 35,    # Дата внесения
            "B": 12,    # ID дефекта
            "C": 18,    # Станция
            "D": 22.5,  # VIN
            "E": 26,    # Модель
            "F": 22,    # Тип дефекта
            "G": 32,    # Область дефекта
            "H": 10,    # Грейд
            "I": 24,    # Ответственный
            "J": 42,    # Комментарий
            "K": 14,    # Устранена
            "L": 20,    # Кто устранил
            "M": 35,    # Дата устранения
            "N": 14,    # Проверена
            "O": 20,    # Кто разрешил
            "P": 24,    # Дата разрешения
            "Q": 24,    # Дата отправки на СГП
            "R": 10,    # Смена
            "S": 70,    # Фото дефекта            
        }

        apply_excel_style(ws, column_widths)

        for row in range(2, ws.max_row + 1):
            ws[f"S{row}"].alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=True
            )
            ws.row_dimensions[row].height = 45

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        file_name = f"defects_{start_datetime.strftime('%Y-%m-%d_%H-%M')}_{end_datetime.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    return render(request, "defects_app/export_defects.html")


@login_required
def defect_photo_detail_view(request, defect_id):
    response = manager_required_view(request)
    if response:
        return response

    defect = get_object_or_404(
        Defekty.objects.select_related(
            "avto",
            "avto__model",
            "mesto",
            "tip",
            "oblast",
            "greyd",
            "otvetstvennyj",
            "smena",
        ).prefetch_related("photos"),
        id=defect_id
    )

    sgp_status = StatusAvto.objects.filter(
        avto=defect.avto,
        status="СГП"
    ).order_by("-data_statusa").first()

    return render(request, "defects_app/defect_photo_detail.html", {
        "defect": defect,
        "sgp_status": sgp_status,
    })

@login_required
def export_created_cars_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")

    response = manager_required_view(request)
    if response:
        return response

    if request.method == "POST":
        start_datetime, end_datetime, redirect_response = get_export_period(request, "export_created_cars")
        if redirect_response:
            return redirect_response

        cars = Avtomobili.objects.filter(
            data_sozdaniya__gte=start_datetime,
            data_sozdaniya__lte=end_datetime
        ).select_related(
            "model"
        ).annotate(
            last_repair_date=Max("defekty__data_ustraneniya"),
            last_approve_date=Max("defekty__data_razresheniya"),
        ).order_by("data_sozdaniya", "model__nazvanie")

        wb = Workbook()
        ws = wb.active
        ws.title = "Созданные машины"

        headers = [
            "Дата создания",
            "VIN",
            "Модель",
            "Кто создал",
            "Когда прошла Бестеневую",
            "Когда прошла Доводку",
            "Когда прошла OKLine",
            "Дата отправки на СГП",
            "Жизненный цикл на линии",
            "Полный жизненный цикл",
        ]

        ws.append(headers)

        for car in cars:
            sgp_status = StatusAvto.objects.filter(
                avto=car,
                status="СГП"
            ).order_by("-data_statusa").first()

            sgp_date = sgp_status.data_statusa if sgp_status else None

            ws.append([
                format_datetime_for_excel(car.data_sozdaniya),
                car.vin,
                car.model.nazvanie if car.model else "",
                car.kto_sozdal or "",
                format_datetime_for_excel(car.data_prohoda_bestenevaya),
                format_datetime_for_excel(car.last_repair_date),
                format_datetime_for_excel(car.last_approve_date),
                format_datetime_for_excel(sgp_date),
                format_duration_for_excel(car.data_sozdaniya, car.data_prohoda_bestenevaya),
                format_duration_for_excel(car.data_sozdaniya, sgp_date),
            ])

        apply_excel_style(ws, {
            "A": 22,
            "B": 22.5,
            "C": 26,
            "D": 20,
            "E": 28,
            "F": 26,
            "G": 26,
            "H": 26,
            "I": 26,
            "J": 26,
        })

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        file_name = f"created_cars_{start_datetime.strftime('%Y-%m-%d_%H-%M')}_{end_datetime.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    return render(request, "defects_app/export_created_cars.html")

@login_required
def production_plans_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")

    response = manager_required_view(request)
    if response:
        return response

    if request.method == "POST":
        date = request.POST.get("date")
        plan_count = request.POST.get("plan_count")
        work_minutes = request.POST.get("work_minutes")

        if not date or not plan_count or not work_minutes:
            messages.error(request, "Заполните дату, план машин и рабочие минуты.")
            return redirect("production_plans")

        plan, created = DailyProductionPlan.objects.get_or_create(
            date=date,
            defaults={
                "plan_count": plan_count,
                "work_minutes": work_minutes,
                "created_by": request.user.username,
                "updated_by": request.user.username,
            }
        )

        if not created:
            plan.plan_count = plan_count
            plan.work_minutes = work_minutes
            plan.updated_by = request.user.username
            plan.save()

        messages.success(request, "План сохранён.")
        return redirect("production_plans")

    plans = DailyProductionPlan.objects.all().order_by("-date")

    return render(request, "defects_app/production_plans.html", {
        "plans": plans
    })

SHIFT_WORK_MINUTES = 450

WORK_INTERVALS = [
    (time(8, 0), time(10, 0)),
    (time(10, 15), time(12, 0)),
    (time(13, 0), time(15, 0)),
    (time(15, 15), time(17, 0)),
]


def datetime_to_work_minutes(value):
    if not value:
        return None

    current_time = value.time()
    total_minutes = 0

    for start, end in WORK_INTERVALS:
        start_dt = datetime.combine(value.date(), start)
        end_dt = datetime.combine(value.date(), end)

        if value >= end_dt:
            total_minutes += int((end_dt - start_dt).total_seconds() / 60)

        elif start_dt <= value < end_dt:
            total_minutes += int((value - start_dt).total_seconds() / 60)
            break

    return total_minutes


def calculate_oee_for_day(day):
    today = timezone.now().date()
    is_future = day > today

    plan = DailyProductionPlan.objects.filter(date=day).first()

    if not plan or not plan.plan_count:
        return {
            "date": day.strftime("%d.%m.%Y"),
            "day": day.strftime("%d"),
            "plan": 0,
            "fact": 0,
            "oee": 0,
            "plan_execution": 0,
            "downtime_minutes": 0,
            "downtime_percent": 0,
            "takt_minutes": "",
            "actual_takt_minutes": "",
            "work_minutes": 0,
            "has_plan": False,
            "is_future": is_future,
            "is_counted": False,
        }

    work_minutes = plan.work_minutes or 450
    takt_minutes = work_minutes / plan.plan_count

    if is_future:
        return {
            "date": day.strftime("%d.%m.%Y"),
            "day": day.strftime("%d"),
            "plan": plan.plan_count,
            "fact": 0,
            "oee": 0,
            "plan_execution": 0,
            "downtime_minutes": 0,
            "downtime_percent": 0,
            "takt_minutes": round(takt_minutes, 2),
            "actual_takt_minutes": "",
            "work_minutes": work_minutes,
            "has_plan": True,
            "is_future": True,
            "is_counted": False,
        }

    start_datetime = datetime.combine(day, time(8, 0))
    end_datetime = datetime.combine(day, time(23, 59))

    cars = Avtomobili.objects.filter(
        data_prohoda_bestenevaya__gte=start_datetime,
        data_prohoda_bestenevaya__lte=end_datetime
    ).exclude(
        data_prohoda_bestenevaya__isnull=True
    ).order_by("data_prohoda_bestenevaya")

    fact_count = cars.count()

    work_marks = []

    for car in cars:
        minutes = datetime_to_work_minutes(car.data_prohoda_bestenevaya)
        if minutes is not None:
            work_marks.append(minutes)

    downtime_minutes = 0

    if fact_count == 0:
        downtime_minutes = work_minutes
    else:
        previous_mark = 0

        for mark in work_marks:
            interval = mark - previous_mark
            delay = interval - takt_minutes

            if delay > 0:
                downtime_minutes += delay

            previous_mark = mark

    downtime_percent = (downtime_minutes / work_minutes) * 100 if work_minutes else 0

    oee = 100 - downtime_percent

    if oee < 0:
        oee = 0

    plan_execution = (fact_count / plan.plan_count) * 100 if plan.plan_count else 0
    actual_takt_minutes = work_minutes / fact_count if fact_count else ""

    return {
        "date": day.strftime("%d.%m.%Y"),
        "day": day.strftime("%d"),
        "plan": plan.plan_count,
        "fact": fact_count,
        "oee": round(oee, 2),
        "plan_execution": round(plan_execution, 2),
        "downtime_minutes": round(downtime_minutes, 1),
        "downtime_percent": round(downtime_percent, 2),
        "takt_minutes": round(takt_minutes, 2),
        "actual_takt_minutes": round(actual_takt_minutes, 2) if actual_takt_minutes != "" else "",
        "work_minutes": work_minutes,
        "has_plan": True,
        "is_future": False,
        "is_counted": True,
    }


@login_required
def qrqc_oee_api_view(request):
    response = manager_required_view(request)
    if response:
        return response

    month = request.GET.get("month")

    if month:
        selected_date = datetime.strptime(month, "%Y-%m").date()
    else:
        selected_date = timezone.now().date()

    first_day = selected_date.replace(day=1)

    if first_day.month == 12:
        next_month = first_day.replace(year=first_day.year + 1, month=1)
    else:
        next_month = first_day.replace(month=first_day.month + 1)

    days_count = (next_month - first_day).days

    days = [
        first_day + timedelta(days=i)
        for i in range(days_count)
    ]

    daily_data = [
        calculate_oee_for_day(day)
        for day in days
    ]

    counted_days = [
        item for item in daily_data
        if item["has_plan"] and item["is_counted"]
    ]

    total_plan = sum(item["plan"] for item in counted_days)
    total_fact = sum(item["fact"] for item in counted_days)
    total_downtime = sum(item["downtime_minutes"] for item in counted_days)
    total_work_minutes = sum(item["work_minutes"] for item in counted_days)

    if counted_days and total_work_minutes:
        month_oee = 100 - (total_downtime / total_work_minutes * 100)

        if month_oee < 0:
            month_oee = 0

        best_day = max(counted_days, key=lambda x: x["oee"])
        worst_day = min(counted_days, key=lambda x: x["oee"])
    else:
        month_oee = 0
        best_day = None
        worst_day = None

    data = {
        "month": selected_date.strftime("%m.%Y"),
        "days": daily_data,
        "summary": {
            "total_plan": total_plan,
            "total_fact": total_fact,
            "difference": total_fact - total_plan,
            "month_oee": round(month_oee, 2),
            "total_downtime": round(total_downtime, 1),
            "downtime_percent": round(
                (total_downtime / total_work_minutes * 100),
                2
            ) if total_work_minutes else 0,
            "best_day": best_day,
            "worst_day": worst_day,
            "total_work_minutes": total_work_minutes,
        }
    }

    return JsonResponse(data)

@login_required
def reports_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")    

    response = manager_required_view(request)
    if response:
        return response

    return render(request, "defects_app/reports.html")


@login_required
def oee_report_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")
    
    response = manager_required_view(request)
    if response:
        return response

    return render(request, "defects_app/oee_report.html")


@login_required
def oee_print_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month")

    if not selected_month:
        selected_month = timezone.now().strftime("%Y-%m")

    return render(request, "defects_app/oee_print.html", {
        "selected_month": selected_month,
    })


@login_required
def qrqc_dashboard_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")
    
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    return render(request, "defects_app/qrqc_dashboard.html", {
        "selected_month": selected_month,
        "selected_day": selected_day,
    })

@login_required
def bestenevaya_timer_api_view(request):
    session_data, redirect_response = require_station_session(request)
    if redirect_response:
        return JsonResponse({"error": "session_expired"}, status=403)

    today = timezone.now().date()
    now = timezone.now()

    plan = DailyProductionPlan.objects.filter(date=today).first()

    if not plan or not plan.plan_count:
        return JsonResponse({
            "has_plan": False,
            "message": "На сегодня не задан план выпуска.",
        })

    takt_minutes = SHIFT_WORK_MINUTES / plan.plan_count

    shift_start = datetime.combine(today, time(8, 0))
    shift_end = datetime.combine(today, time(17, 0))

    fact_count = Avtomobili.objects.filter(
        data_prohoda_bestenevaya__gte=shift_start,
        data_prohoda_bestenevaya__lte=shift_end
    ).count()

    current_work_minutes = datetime_to_work_minutes(now)

    expected_count = int(current_work_minutes // takt_minutes) if current_work_minutes else 0
    next_target_minutes = (fact_count + 1) * takt_minutes
    minutes_to_next = next_target_minutes - current_work_minutes

    lag_count = max(expected_count - fact_count, 0)

    if minutes_to_next <= 0:
        status = "red"
    elif minutes_to_next <= 2:
        status = "yellow"
    else:
        status = "green"

    return JsonResponse({
        "has_plan": True,
        "plan": plan.plan_count,
        "fact": fact_count,
        "expected": expected_count,
        "lag_count": lag_count,
        "takt_minutes": round(takt_minutes, 2),
        "minutes_to_next": round(minutes_to_next, 1),
        "status": status,
    })

@login_required
def qrqc_dphu_api_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    year, month = map(int, selected_month.split("-"))
    days_in_month = monthrange(year, month)[1]

    day_date = datetime.strptime(selected_day, "%Y-%m-%d").date()

    day_start = datetime.combine(day_date, time.min)
    day_end = datetime.combine(day_date, time.max)

    month_start = datetime.combine(date(year, month, 1), time.min)
    month_end = datetime.combine(date(year, month, days_in_month), time.max)

    responsibles = list(
        Otvetstvennye.objects.filter(aktiven=True).order_by("id")
    )

    responsible_names = [item.nazvanie for item in responsibles]

    # ---------- ДЕНЬ ----------
    day_defects_raw = (
        Defekty.objects.filter(
            data__gte=day_start,
            data__lte=day_end,
            otvetstvennyj__isnull=False,
        )
        .values("otvetstvennyj__nazvanie")
        .annotate(count=Count("id"))
    )

    day_counts = {name: 0 for name in responsible_names}

    for item in day_defects_raw:
        day_counts[item["otvetstvennyj__nazvanie"]] = item["count"]

    day_dphu = {
        name: round(day_counts[name] / 100, 4)
        for name in responsible_names
    }

    day_total_defects = sum(day_counts.values())

    day_cars_fact = Avtomobili.objects.filter(
        data_prohoda_bestenevaya__gte=day_start,
        data_prohoda_bestenevaya__lte=day_end,
    ).count()

    day_dphu_total = round(day_total_defects / 100, 4)

    # ---------- МЕСЯЦ ----------
    month_report = []

    month_total_defects = 0
    month_total_cars = 0

    for day_number in range(1, days_in_month + 1):
        current_date = date(year, month, day_number)

        current_start = datetime.combine(current_date, time.min)
        current_end = datetime.combine(current_date, time.max)

        cars_fact = Avtomobili.objects.filter(
            data_prohoda_bestenevaya__gte=current_start,
            data_prohoda_bestenevaya__lte=current_end,
        ).count()

        month_total_cars += cars_fact

        defects_raw = (
            Defekty.objects.filter(
                data__gte=current_start,
                data__lte=current_end,
                otvetstvennyj__isnull=False,
            )
            .values("otvetstvennyj__nazvanie")
            .annotate(count=Count("id"))
        )

        counts = {name: 0 for name in responsible_names}

        for item in defects_raw:
            counts[item["otvetstvennyj__nazvanie"]] = item["count"]

        total_defects_for_day = sum(counts.values())
        month_total_defects += total_defects_for_day

        if cars_fact > 0:
            dphu = {
                name: round(counts[name] / cars_fact, 4)
                for name in responsible_names
            }
            dphu_total = round(total_defects_for_day / cars_fact, 4)
        else:
            dphu = {
                name: 0
                for name in responsible_names
            }
            dphu_total = 0

        month_report.append({
            "day": f"{day_number:02}",
            "date": current_date.strftime("%Y-%m-%d"),
            "cars_fact": cars_fact,
            "counts": counts,
            "dphu": dphu,
            "total_defects": total_defects_for_day,
            "dphu_total": dphu_total,
        })

    if month_total_cars > 0:
        month_dphu_total = round(month_total_defects / month_total_cars, 4)
    else:
        month_dphu_total = 0

    return JsonResponse({
        "month": selected_month,
        "day": selected_day,
        "responsibles": responsible_names,

        "day_report": {
            "cars_fact": day_cars_fact,
            "base_cars": 100,
            "counts": day_counts,
            "dphu": day_dphu,
            "total_defects": day_total_defects,
            "dphu_total": day_dphu_total,
        },

        "month_report": month_report,

        "month_summary": {
            "total_cars": month_total_cars,
            "total_defects": month_total_defects,
            "dphu_total": month_dphu_total,
        }
    })

@login_required
def dphu_report_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    return render(request, "defects_app/dphu_report.html", {
        "selected_month": selected_month,
        "selected_day": selected_day,
    })

# Диаграмма топ дефектов за смену
@login_required
def qrqc_top_defects_api_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    day_date = datetime.strptime(selected_day, "%Y-%m-%d").date()
    day_start = datetime.combine(day_date, time(8, 0))
    day_end = datetime.combine(day_date, time(17, 0))

    result = {}

    for grade in ["V1", "V2"]:
        defects = (
            Defekty.objects.filter(
                data__gte=day_start,
                data__lte=day_end,
                greyd__nazvanie=grade,
                tip__isnull=False,
            )
            .values("greyd__nazvanie", "tip__nazvanie")
            .annotate(count=Count("id"))
            .order_by("-count", "tip__nazvanie")[:10]
        )

        result[grade] = [
            {
                "grade": item["greyd__nazvanie"],
                "type": item["tip__nazvanie"],
                "count": item["count"],
            }
            for item in defects
        ]

    return JsonResponse({
        "day": selected_day,
        "items": result,
    })

@login_required
def top_defects_report_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")

    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    return render(request, "defects_app/top_defects_report.html", {
        "selected_month": selected_month,
        "selected_day": selected_day,
    })


@login_required
def top_defects_api_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    year, month = map(int, selected_month.split("-"))
    days_in_month = monthrange(year, month)[1]

    day_date = datetime.strptime(selected_day, "%Y-%m-%d").date()

    day_start = datetime.combine(day_date, time.min)
    day_end = datetime.combine(day_date, time.max)

    month_start = datetime.combine(date(year, month, 1), time.min)
    month_end = datetime.combine(date(year, month, days_in_month), time.max)

    def get_top_defects(start_datetime, end_datetime, grade_names):
        return list(
            Defekty.objects.filter(
                data__gte=start_datetime,
                data__lte=end_datetime,
                greyd__nazvanie__in=grade_names,
            )
            .values(
                "tip__nazvanie",
                "oblast__nazvanie",
            )
            .annotate(count=Count("id"))
            .order_by("-count", "tip__nazvanie", "oblast__nazvanie")[:10]
        )

    day_v1 = get_top_defects(day_start, day_end, ["V1", "В1", "v1", "в1"])
    day_v2 = get_top_defects(day_start, day_end, ["V2", "В2", "v2", "в2"])

    month_v1 = get_top_defects(month_start, month_end, ["V1", "В1", "v1", "в1"])
    month_v2 = get_top_defects(month_start, month_end, ["V2", "В2", "v2", "в2"])

    return JsonResponse({
        "month": selected_month,
        "selected_day": selected_day,

        "day_report": {
            "v1": day_v1,
            "v2": day_v2,
            "total_v1": sum(item["count"] for item in day_v1),
            "total_v2": sum(item["count"] for item in day_v2),
            "total": sum(item["count"] for item in day_v1) + sum(item["count"] for item in day_v2),
        },

        "month_report": {
            "v1": month_v1,
            "v2": month_v2,
            "total_v1": sum(item["count"] for item in month_v1),
            "total_v2": sum(item["count"] for item in month_v2),
            "total": sum(item["count"] for item in month_v1) + sum(item["count"] for item in month_v2),
        }
    })

# Печать QRQC отчета
@login_required
def qrqc_print_view(request):
    response = manager_required_view(request)
    if response:
        return response

    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    selected_day = request.GET.get("day") or timezone.now().strftime("%Y-%m-%d")

    return render(request, "defects_app/qrqc_print.html", {
        "selected_month": selected_month,
        "selected_day": selected_day,
    })

# График по моделям и количесву машин на СНП
@login_required
def qrqc_snp_api_view(request):
    response = manager_required_view(request)
    if response:
        return response
    selected_day = request.GET.get("day")
    selected_month = request.GET.get("month") or timezone.now().strftime("%Y-%m")

    latest_status = StatusAvto.objects.filter(
        avto=OuterRef("pk")
    ).order_by("-data_statusa")

    cars = Avtomobili.objects.annotate(
        last_status=Subquery(latest_status.values("status")[:1]),
        last_status_date=Subquery(latest_status.values("data_statusa")[:1]),
    ).filter(
        last_status="СНП"
    ).select_related("model").order_by("last_status_date")

    # ---------- ГРАФИК ПО МОДЕЛЯМ ----------
    model_counts = {}

    for car in cars:
        model_name = car.model.nazvanie if car.model else "Без модели"

        if model_name not in model_counts:
            model_counts[model_name] = 0

        model_counts[model_name] += 1

    models_chart = [
        {
            "model": model,
            "count": count
        }
        for model, count in model_counts.items()
    ]

    # ---------- ТОП 10 СТАРЫХ МАШИН НА СНП ----------
    today = timezone.now()

    oldest_cars = []

    for car in cars[:10]:
        days_on_snp = 0

        if car.last_status_date:
            days_on_snp = (today - car.last_status_date).days

        comments_qs = SnpDefectComment.objects.filter(
            avto=car
        ).select_related(
            "defect",
            "defect__tip",
            "defect__oblast"
        ).order_by("id")

        comments = []

        for item in comments_qs:
            defect_text = ""

            if item.defect:
                tip = item.defect.tip.nazvanie if item.defect.tip else ""
                oblast = item.defect.oblast.nazvanie if item.defect.oblast else ""

                defect_text = f"{tip} / {oblast}".strip(" /")

            if defect_text:
                comments.append(f"{defect_text}: {item.comment}")
            else:
                comments.append(item.comment)

        oldest_cars.append({
            "vin": car.vin,
            "model": car.model.nazvanie if car.model else "",
            "days_on_snp": days_on_snp,
            "snp_date": car.last_status_date.strftime("%d.%m.%Y %H:%M") if car.last_status_date else "",
            "comments": comments,
        })

    latest_day_snp_cars = []
    if selected_day:
        try:
            selected_date = datetime.strptime(selected_day, "%Y-%m-%d").date()
            day_start = datetime.combine(selected_date, time.min)
            day_end = datetime.combine(selected_date, time.max)

            day_snp_statuses = (
                StatusAvto.objects.filter(
                    status="СНП",
                    data_statusa__gte=day_start,
                    data_statusa__lte=day_end,
                )
                .select_related("avto", "avto__model")
                .order_by("-data_statusa")[:10]
            )

            for status_item in day_snp_statuses:
                car = status_item.avto

                comments_qs = SnpDefectComment.objects.filter(
                    avto=car
                ).select_related(
                    "defect",
                    "defect__tip",
                    "defect__oblast"
                ).order_by("id")

                comments = []

                for item in comments_qs:
                    defect_text = ""

                    if item.defect:
                        tip = item.defect.tip.nazvanie if item.defect.tip else ""
                        oblast = item.defect.oblast.nazvanie if item.defect.oblast else ""

                        defect_text = f"{tip} / {oblast}".strip(" /")

                    if defect_text:
                        comments.append(f"{defect_text}: {item.comment}")
                    else:
                        comments.append(item.comment)

                latest_day_snp_cars.append({
                    "vin": car.vin if car else "",
                    "model": car.model.nazvanie if car and car.model else "",
                    "when": status_item.data_statusa.strftime("%H:%M"),
                    "comments": comments,
                })
        except ValueError:
            latest_day_snp_cars = []

    snp_in_month_chart = []
    avg_snp_per_day_chart = []
    avg_snp_per_month_chart = []

    try:
        month_date = datetime.strptime(selected_month, "%Y-%m").date()
        month_start = datetime.combine(month_date.replace(day=1), time.min)
        days_in_month = monthrange(month_date.year, month_date.month)[1]
        month_end = datetime.combine(date(month_date.year, month_date.month, days_in_month), time.max)

        # 1) Сколько машин на СНП по дням выбранного месяца (остаток на конец дня)
        for day_number in range(1, days_in_month + 1):
            current_day = date(month_date.year, month_date.month, day_number)
            current_day_end = datetime.combine(current_day, time.max)

            latest_status_to_day = StatusAvto.objects.filter(
                avto=OuterRef("pk"),
                data_statusa__lte=current_day_end
            ).order_by("-data_statusa")

            day_snp_count = Avtomobili.objects.annotate(
                last_status=Subquery(latest_status_to_day.values("status")[:1]),
            ).filter(
                last_status="СНП"
            ).count()

            snp_in_month_chart.append({
                "day": f"{day_number:02}",
                "count": day_snp_count,
            })

        # 2) Сколько машин ушло на СНП по дням выбранного месяца
        for day_number in range(1, days_in_month + 1):
            current_day = date(month_date.year, month_date.month, day_number)
            day_start = datetime.combine(current_day, time.min)
            day_end = datetime.combine(current_day, time.max)

            day_snp_out_count = StatusAvto.objects.filter(
                status="СНП",
                data_statusa__gte=day_start,
                data_statusa__lte=day_end,
            ).values("avto_id").distinct().count()

            avg_snp_per_day_chart.append({
                "day": f"{day_number:02}",
                "count": day_snp_out_count,
            })

        # 3) Среднее число ушедших на СНП по месяцам выбранного года
        for month_number in range(1, 13):
            month_days = monthrange(month_date.year, month_number)[1]
            current_month_start = datetime.combine(date(month_date.year, month_number, 1), time.min)
            current_month_end = datetime.combine(date(month_date.year, month_number, month_days), time.max)

            month_snp_out_count = StatusAvto.objects.filter(
                status="СНП",
                data_statusa__gte=current_month_start,
                data_statusa__lte=current_month_end,
            ).values("avto_id").distinct().count()

            month_avg = round(month_snp_out_count / month_days, 2) if month_days else 0

            avg_snp_per_month_chart.append({
                "month": f"{month_number:02}",
                "avg_count": month_avg,
            })
    except ValueError:
        pass

    return JsonResponse({
        "models_chart": models_chart,
        "oldest_cars": oldest_cars,
        "latest_day_snp_cars": latest_day_snp_cars,
        "snp_in_month_chart": snp_in_month_chart,
        "avg_snp_per_day_chart": avg_snp_per_day_chart,
        "avg_snp_per_month_chart": avg_snp_per_month_chart,
    })

PRODUCTION_INTERVALS = [
    ("8:00", "9:00"),
    ("9:00", "10:00"),
    ("10:15", "11:00"),
    ("11:00", "12:00"),
    ("13:00", "14:00"),
    ("14:00", "15:00"),
    ("15:15", "16:00"),
    ("16:00", "17:00"),
    ("17:15", "18:00"),
    ("18:00", "19:00"),
    ("19:15", "20:00"),
]


def time_text_to_time(value):
    return datetime.strptime(value, "%H:%M").time()


def get_minutes_between(start_text, end_text):
    start = datetime.strptime(start_text, "%H:%M")
    end = datetime.strptime(end_text, "%H:%M")
    return int((end - start).total_seconds() / 60)


def add_minutes_to_time(start_text, minutes):
    start = datetime.strptime(start_text, "%H:%M")
    end = start + timedelta(minutes=minutes)
    return end.strftime("%H:%M")


def format_minutes_to_hhmm(minutes):
    if not minutes:
        return ""

    minutes = int(round(minutes))
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}:{mins:02d}"


def get_active_intervals(work_minutes):
    result = []
    total = 0

    for start, end in PRODUCTION_INTERVALS:
        interval_minutes = get_minutes_between(start, end)

        if total >= work_minutes:
            break

        if total + interval_minutes <= work_minutes:
            result.append((start, end, interval_minutes))
            total += interval_minutes
        else:
            remaining = work_minutes - total
            if remaining > 0:
                new_end = add_minutes_to_time(start, remaining)
                result.append((start, new_end, remaining))
            break

    return result


def distribute_plan_by_intervals(total_plan, active_intervals):
    if not total_plan or not active_intervals:
        return [0 for _ in active_intervals]

    total_minutes = sum(item[2] for item in active_intervals)

    raw_values = [
        total_plan * interval_minutes / total_minutes
        for _, _, interval_minutes in active_intervals
    ]

    rounded = [round(value) for value in raw_values]
    difference = total_plan - sum(rounded)

    while difference != 0:
        if difference > 0:
            index = rounded.index(min(rounded))
            rounded[index] += 1
            difference -= 1
        else:
            indexes = [i for i, value in enumerate(rounded) if value > 0]
            if not indexes:
                break
            index = indexes[-1]
            rounded[index] -= 1
            difference += 1

    return rounded


@login_required
def export_qrqc_production_view(request):
    response = manager_required_view(request)
    if response:
        return response

    month = request.GET.get("month")

    if month:
        selected_date = datetime.strptime(month, "%Y-%m").date()
    else:
        selected_date = timezone.localtime().date()

    today = datetime.now().date()

    if selected_date.year == today.year and selected_date.month == today.month:
        report_date = today
    else:
        report_date = selected_date.replace(day=1)

    plan = DailyProductionPlan.objects.filter(date=report_date).first()

    plan_count = plan.plan_count if plan else 0
    work_minutes = plan.work_minutes if plan and plan.work_minutes else 450

    active_intervals = get_active_intervals(work_minutes)
    interval_plan = distribute_plan_by_intervals(plan_count, active_intervals)

    day_start = timezone.make_aware(datetime.combine(report_date, time(0, 0)))
    day_end = timezone.make_aware(datetime.combine(report_date, time(23, 59, 59)))

    cars = Avtomobili.objects.filter(
        data_prohoda_bestenevaya__gte=day_start,
        data_prohoda_bestenevaya__lte=day_end,
    ).exclude(
        data_prohoda_bestenevaya__isnull=True
    ).select_related("model")

    wb = Workbook()
    ws = wb.active
    ws.title = "QRQC"

    black_fill = PatternFill("solid", fgColor="000000")
    blue_fill = PatternFill("solid", fgColor="DDEBF7")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=18)
    header_font = Font(bold=True, color="FFFFFF", size=9)
    bold_font = Font(bold=True)
    white_bold_font = Font(bold=True, color="FFFFFF")
    red_bold_font = Font(bold=True, color="FF0000")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A1:L1")
    ws["A1"] = "Отчет производства линия EVOLUTE i-SPACE"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["N1"] = report_date.strftime("%d.%m.%Y")
    ws["N1"].font = Font(bold=True, size=12)
    ws["N1"].alignment = center

    headers = [
        "Интервал", "Начало", "Конец", "План", "Факт",
        "Разница", "Отклонение", "Время простоя",
        "Участок", "Ответственный", "Причина простоя", "Контрмеры",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = black_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    total_fact = 0
    total_downtime = 0
    now = datetime.now()

    for index, interval in enumerate(active_intervals, start=1):
        row = index + 3
        start_text, end_text, interval_minutes = interval

        start_dt = datetime.combine(report_date, time_text_to_time(start_text))
        end_dt = datetime.combine(report_date, time_text_to_time(end_text))

        fact_start_dt = start_dt
        fact_end_dt = end_dt

        if start_text == "10:15":
            fact_start_dt = datetime.combine(report_date, time(10, 0))

        if start_text == "13:00":
            fact_start_dt = datetime.combine(report_date, time(12, 0))

        if start_text == "15:15":
            fact_start_dt = datetime.combine(report_date, time(15, 0))

        if start_text == "17:15":
            fact_start_dt = datetime.combine(report_date, time(17, 0))

        if start_text == "19:15":
            fact_start_dt = datetime.combine(report_date, time(19, 0))

        fact = cars.filter(
            data_prohoda_bestenevaya__gte=fact_start_dt,
            data_prohoda_bestenevaya__lt=fact_end_dt
        ).count()

        plan_value = interval_plan[index - 1]
        difference = fact - plan_value

        is_not_finished = report_date == today and end_dt > now

        if is_not_finished:
            downtime = 0
            difference_value = ""
            deviation_value = ""
        else:
            downtime = abs(difference) * (work_minutes / plan_count) if difference < 0 and plan_count else 0
            difference_value = difference if difference else ""
            deviation_value = difference if difference else ""

        total_fact += fact
        total_downtime += downtime

        values = [
            index,
            start_text,
            end_text,
            plan_value,
            fact if fact else "",
            difference_value,
            deviation_value,
            format_minutes_to_hhmm(downtime),
            "",
            "",
            "",
            "",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = center
            cell.fill = blue_fill if index % 2 == 0 else white_fill

            if col in [6, 7] and isinstance(value, int) and value < 0:
                cell.font = red_bold_font
            elif col in [1, 2, 3, 4, 5]:
                cell.font = bold_font

    total_row = len(active_intervals) + 4
    total_difference = total_fact - plan_count

    total_values = [
        len(active_intervals),
        active_intervals[0][0] if active_intervals else "",
        active_intervals[-1][1] if active_intervals else "",
        plan_count,
        total_fact,
        total_difference,
        "",
        format_minutes_to_hhmm(total_downtime),
        "",
        "",
        "",
        "",
    ]

    for col, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col)
        cell.value = value
        cell.fill = black_fill
        cell.font = red_bold_font if col in [5, 6] else white_bold_font
        cell.border = border
        cell.alignment = center

    model_block_start_col = 14

    ws.cell(row=3, column=model_block_start_col).value = "Модель"
    ws.cell(row=3, column=model_block_start_col + 1).value = "Количество авто"

    for col in [model_block_start_col, model_block_start_col + 1]:
        cell = ws.cell(row=3, column=col)
        cell.fill = black_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ispace_count = 0
    ijoy_count = 0
    free_count = 0

    for car in cars:
        model_name = car.model.nazvanie.upper() if car.model else ""

        if "I-SPACE" in model_name:
            ispace_count += 1
        elif "I-JOY" in model_name or "IJOY" in model_name:
            ijoy_count += 1
        elif "FREE" in model_name:
            free_count += 1

    model_total = ispace_count + ijoy_count + free_count

    model_rows = [
        ("I-SPACE", ispace_count),
        ("I-JOY", ijoy_count),
        ("FREE", free_count),
    ]

    for i, row_data in enumerate(model_rows, start=4):
        for j, value in enumerate(row_data, start=model_block_start_col):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = border
            cell.alignment = center
            cell.font = bold_font
            cell.fill = white_fill if i % 2 == 0 else blue_fill

    summary_rows = [
        ("Общий итог", model_total),
        ("Разница", model_total - plan_count),
    ]

    for i, row_data in enumerate(summary_rows, start=7):
        for j, value in enumerate(row_data, start=model_block_start_col):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.fill = black_fill
            cell.border = border
            cell.alignment = center
            cell.font = red_bold_font if row_data[0] == "Разница" and j == model_block_start_col + 1 else white_bold_font

    ws.merge_cells(start_row=10, start_column=model_block_start_col, end_row=10, end_column=model_block_start_col + 1)
    comment_title = ws.cell(row=10, column=model_block_start_col)
    comment_title.value = "Комментарии"
    comment_title.fill = black_fill
    comment_title.font = white_bold_font
    comment_title.alignment = center
    comment_title.border = border

    ws.merge_cells(start_row=11, start_column=model_block_start_col, end_row=total_row, end_column=model_block_start_col + 1)
    comment_area = ws.cell(row=11, column=model_block_start_col)
    comment_area.fill = blue_fill
    comment_area.border = border
    comment_area.alignment = left

    widths = {
        "A": 8, "B": 10, "C": 10, "D": 8, "E": 8,
        "F": 9, "G": 11, "H": 12, "I": 12, "J": 16,
        "K": 30, "L": 32, "M": 4, "N": 14, "O": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(3, total_row + 1):
        ws.row_dimensions[row].height = 30

    ws.row_dimensions[1].height = 28

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    file_name = f"qrqc_production_{report_date.strftime('%Y-%m-%d')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response

# внесение ВИН-номеров и их префиксов
@login_required
def vin_prefixes_api_view(request):
    prefixes = VinPrefix.objects.filter(
        is_active=True,
        show_in_select=True
    ).select_related("model").order_by("model__nazvanie", "prefix")

    data = []

    for item in prefixes:
        data.append({
            "model": item.model.nazvanie,
            "prefix": item.prefix,
        })

    return JsonResponse(data, safe=False)


@login_required
def vin_model_api_view(request):
    vin = request.GET.get("vin", "").strip().upper()

    if len(vin) != 17:
        return JsonResponse({
            "found": False,
            "model": None,
        })

    # Плановые VIN используем ТОЛЬКО для FREE,
    # потому что у FREE / FREE DKD / FREE MKD одинаковый префикс
    if vin.startswith("EDAVGC3B0TL"):
        plan_vin = PlanovyeVin.objects.filter(vin=vin).first()

        if plan_vin and plan_vin.model:
            return JsonResponse({
                "found": True,
                "source": "plan",
                "model": plan_vin.model.strip(),
            })

    # Все остальные модели определяем ТОЛЬКО по префиксу
    prefixes = VinPrefix.objects.filter(
        is_active=True,
        prefix__isnull=False
    ).select_related("model").order_by("-prefix")

    for item in prefixes:
        if vin.startswith(item.prefix):
            return JsonResponse({
                "found": True,
                "source": "prefix",
                "model": item.model.nazvanie,
            })

    return JsonResponse({
        "found": False,
        "model": None,
    })

@login_required
def upload_old_snp_cars_view(request):
    response = manager_required_view(request)
    if response:
        return response

    result = None

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Выберите Excel-файл.")
            return redirect("upload_old_snp_cars")

        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        created_cars = 0
        updated_existing = 0
        created_defects = 0
        created_statuses = 0
        errors = []

        default_tip, _ = Tipy.objects.get_or_create(nazvanie="Прочее")
        default_oblast, _ = Oblasti.objects.get_or_create(nazvanie="СНП")
        default_greyd, _ = Greydy.objects.get_or_create(nazvanie="V2")
        default_mesto = Mesta.objects.filter(id=2).first() or Mesta.objects.first()
        default_smena = Smeny.objects.first()

        if not default_mesto or not default_smena:
            messages.error(request, "В базе нет места или смены. Сначала проверь справочники.")
            return redirect("upload_old_snp_cars")

        headers = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value:
                headers[str(value).strip().lower()] = col

        vin_col = headers.get("vin")
        date_col = headers.get("дата снп")
        comment_col = headers.get("комментарий")
        tip_col = headers.get("тип")
        oblast_col = headers.get("область")
        greyd_col = headers.get("грейд")

        if not vin_col or not date_col:
            messages.error(request, "В файле обязательно должны быть колонки VIN и Дата СНП.")
            return redirect("upload_old_snp_cars")

        for row in range(2, ws.max_row + 1):
            vin = ws.cell(row=row, column=vin_col).value
            snp_date = ws.cell(row=row, column=date_col).value

            if not vin:
                continue

            vin = str(vin).strip().upper()

            if len(vin) != 17:
                errors.append(f"Строка {row}: VIN {vin} не 17 символов.")
                continue

            if not snp_date:
                errors.append(f"Строка {row}: не указана дата СНП.")
                continue

            if isinstance(snp_date, datetime):
                snp_datetime = snp_date
            else:
                try:
                    snp_datetime = datetime.strptime(str(snp_date).strip(), "%d.%m.%Y")
                except ValueError:
                    try:
                        snp_datetime = datetime.strptime(str(snp_date).strip(), "%Y-%m-%d")
                    except ValueError:
                        errors.append(f"Строка {row}: не удалось прочитать дату СНП.")
                        continue

            bestenevaya_datetime = snp_datetime - timedelta(days=1)

            model_obj = None

            plan_vin = PlanovyeVin.objects.filter(vin=vin).first()
            if plan_vin:
                model_obj, _ = Modeli.objects.get_or_create(
                    nazvanie=plan_vin.model.strip()
                )

            if not model_obj:
                prefixes = VinPrefix.objects.filter(
                    is_active=True,
                    prefix__isnull=False
                ).select_related("model").order_by("-prefix")

                for item in prefixes:
                    if vin.startswith(item.prefix):
                        model_obj = item.model
                        break

            if not model_obj:
                errors.append(f"Строка {row}: не удалось определить модель для VIN {vin}.")
                continue

            existing_car = Avtomobili.objects.filter(vin=vin).first()

            if existing_car:
                car = existing_car

                car.model = model_obj
                car.kto_sozdal = car.kto_sozdal or request.user.username
                car.data_sozdaniya = bestenevaya_datetime
                car.proshla_bestenevaya = True
                car.data_prohoda_bestenevaya = bestenevaya_datetime
                car.save(update_fields=[
                    "model",
                    "kto_sozdal",
                    "data_sozdaniya",
                    "proshla_bestenevaya",
                    "data_prohoda_bestenevaya",
                ])

                Defekty.objects.filter(
                    avto=car,
                    tip=default_tip,
                    oblast=default_oblast
                ).delete()

                StatusAvto.objects.filter(
                    avto=car,
                    status="СНП"
                ).delete()

                updated_existing += 1

            else:
                car = Avtomobili.objects.create(
                    vin=vin,
                    model=model_obj,
                    kto_sozdal=request.user.username,
                    data_sozdaniya=bestenevaya_datetime,
                    proshla_bestenevaya=True,
                    data_prohoda_bestenevaya=bestenevaya_datetime,
                )

                created_cars += 1

            tip_text = ws.cell(row=row, column=tip_col).value if tip_col else ""
            oblast_text = ws.cell(row=row, column=oblast_col).value if oblast_col else ""
            greyd_text = ws.cell(row=row, column=greyd_col).value if greyd_col else ""
            comment_text = ws.cell(row=row, column=comment_col).value if comment_col else ""

            comment_parts = []

            if tip_text:
                comment_parts.append(f"Тип из старого файла: {tip_text}")
            if oblast_text:
                comment_parts.append(f"Область из старого файла: {oblast_text}")
            if greyd_text:
                comment_parts.append(f"Грейд из старого файла: {greyd_text}")
            if comment_text:
                comment_parts.append(f"Комментарий: {comment_text}")

            final_comment = "\n".join(comment_parts) if comment_parts else "Старый дефект из файла СНП"

            defect = Defekty.objects.create(
                avto=car,
                smena=default_smena,
                mesto=default_mesto,
                tip=default_tip,
                oblast=default_oblast,
                greyd=default_greyd,
                kommentarij=final_comment,
                kto_sozdal=request.user.username,
            )
            created_defects += 1

            SnpDefectComment.objects.create(
                defect=defect,
                avto=car,
                comment=final_comment,
                kto_sozdal=request.user.username,
                data_sozdaniya=snp_datetime,
            )

            StatusAvto.objects.create(
                avto=car,
                status="СНП",
                data_statusa=snp_datetime,
                kto_izmenil=request.user.username
            )
            created_statuses += 1

        result = {
            "created_cars": created_cars,
            "skipped_existing": updated_existing,
            "created_defects": created_defects,
            "created_statuses": created_statuses,
            "errors": errors,
        }

        messages.success(request, "Загрузка завершена.")

    return render(request, "defects_app/upload_old_snp_cars.html", {
        "result": result,
        "is_manager": is_manager_user(request.user),
    })


@login_required
def export_snp_cars_view(request):
    response = manager_required_view(request)
    if response:
        return response

    if request.method == "POST":
        start_datetime, end_datetime, redirect_response = get_export_period(request, "export_snp_cars")
        if redirect_response:
            return redirect_response

        snp_comments = SnpDefectComment.objects.filter(
            data_sozdaniya__gte=start_datetime,
            data_sozdaniya__lte=end_datetime
        ).select_related(
            "avto",
            "avto__model",
            "defect",
            "defect__tip",
            "defect__oblast",
            "defect__greyd",
            "defect__otvetstvennyj",
        ).order_by("data_sozdaniya", "avto__vin")

        wb = Workbook()
        ws = wb.active
        ws.title = "СНП"

        headers = [
            "Дата отправки на СНП",
            "VIN",
            "Модель",
            "Тип дефекта",
            "Область",
            "Грейд",
            "Ответственный",
            "Комментарий дефекта",
            "Комментарий причины СНП",
            "Кто отправил/создал комментарий",
        ]

        ws.append(headers)

        for item in snp_comments:
            defect = item.defect
            car = item.avto

            ws.append([
                format_datetime_for_excel(item.data_sozdaniya),
                car.vin if car else "",
                car.model.nazvanie if car and car.model else "",
                defect.tip.nazvanie if defect and defect.tip else "",
                defect.oblast.nazvanie if defect and defect.oblast else "",
                defect.greyd.nazvanie if defect and defect.greyd else "",
                defect.otvetstvennyj.nazvanie if defect and defect.otvetstvennyj else "",
                defect.kommentarij if defect and defect.kommentarij else "",
                item.comment or "",
                item.kto_sozdal or "",
            ])

        column_widths = {
            "A": 26,
            "B": 22.5,
            "C": 28,
            "D": 28,
            "E": 34,
            "F": 12,
            "G": 28,
            "H": 50,
            "I": 55,
            "J": 28,
        }

        apply_excel_style(ws, column_widths)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        file_name = f"snp_cars_{start_datetime.strftime('%Y-%m-%d_%H-%M')}_{end_datetime.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    return render(request, "defects_app/export_snp_cars.html")

# Большая выгрузка по машине
@login_required
def export_full_cars_view(request):
    response = manager_required_view(request)
    if response:
        return response

    if request.method == "POST":
        period = request.POST.get("period")

        if period == "all":
            start_datetime = None
            end_datetime = None

            cars = Avtomobili.objects.select_related("model").order_by(
                "data_sozdaniya",
                "data_prohoda_bestenevaya",
                "vin"
            )
        else:
            start_datetime, end_datetime, redirect_response = get_export_period(request, "export_full_cars")
            if redirect_response:
                return redirect_response

            cars = Avtomobili.objects.filter(
                data_sozdaniya__gte=start_datetime,
                data_sozdaniya__lte=end_datetime
            ).select_related("model").order_by(
                "data_sozdaniya",
                "data_prohoda_bestenevaya",
                "vin"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Машины"

        headers = [
            "№ п/п",
            "Дата прихода кузова",
            "Номер лота",
            "Модель",
            "VIN РФ",
            "VIN присвоенный поставщиком",
            "Цвет кузова",
            "Цвет салона",
            "Номер сварки",
            "Комплектация",
            "Дата постановки на конвейер",
            "Дата ухода с конвейера",
            "Номер переднего электродвигателя",
            "Номер заднего электродвигателя",
            "Номер двигателя внутреннего сгорания",
            "Дата активации в ГАИС ЭРА-ГЛОНАСС",
            "ГЛОНАСС полностью",
            "Серийный номер ЭРА ГЛОНАСС",
            "ICCID ЭРА ГЛОНАСС",
            "IMEI ЭРА ГЛОНАСС",
            "ID телематики",
            "ОТТС",
            "Номер батареи",
            "Когда привязали батарею",
            "Дата прохождения участка доводки",
            "Дата прохождения дорожных испытаний",
            "Дата возврата на доводку",
            "Дата прихода на СНП",
            "Дата ухода на СГП",
            "Дата прихода на СВХ",
            "Отгружен",
            "Дата отгрузки",
        ]

        ws.append(headers)

        for index, car in enumerate(cars, start=1):
            plan = PlanovyeVin.objects.filter(vin=car.vin).first()

            first_repair = Defekty.objects.filter(
                avto=car,
                ustraneno=True,
                data_ustraneniya__isnull=False
            ).order_by("data_ustraneniya").first()

            snp_status = StatusAvto.objects.filter(
                avto=car,
                status="СНП"
            ).order_by("data_statusa").first()

            sgp_status = StatusAvto.objects.filter(
                avto=car,
                status="СГП"
            ).order_by("data_statusa").first()

            return_to_dovodka = None
            if snp_status:
                return_to_dovodka = Defekty.objects.filter(
                    avto=car,
                    ustraneno=True,
                    data_ustraneniya__gt=snp_status.data_statusa
                ).order_by("data_ustraneniya").first()

            ws.append([
                index,
                "",

                plan.nomer_lota if plan else "",
                car.model.nazvanie if car.model else "",
                car.vin,
                "",

                plan.cvet_kuzova if plan else "",
                plan.cvet_salona if plan else "",
                "",
                plan.komplektaciya if plan else "",

                format_datetime_for_excel(car.data_sozdaniya),
                format_datetime_for_excel(car.data_prohoda_bestenevaya),

                car.perednij_dvigatel or "",
                car.zadnij_dvigatel or "",
                "",

                "",
                car.glonass or "",
                car.glonass_sn or "",
                car.glonass_iccid or "",
                car.glonass_imei or "",
                car.telematika or "",

                plan.otts if plan else "",
                car.batareya or "",
                format_datetime_for_excel(car.batareya_kogda),

                format_datetime_for_excel(first_repair.data_ustraneniya) if first_repair else "",
                "",

                format_datetime_for_excel(return_to_dovodka.data_ustraneniya) if return_to_dovodka else "",
                format_datetime_for_excel(snp_status.data_statusa) if snp_status else "",
                format_datetime_for_excel(sgp_status.data_statusa) if sgp_status else "",

                "",
                "Нет",
                "",
            ])

        apply_excel_style(ws, {
            "A": 8,
            "B": 24,
            "C": 18,
            "D": 28,
            "E": 22,
            "F": 28,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 28,
            "K": 26,
            "L": 26,
            "M": 30,
            "N": 30,
            "O": 30,
            "P": 30,
            "Q": 45,
            "R": 28,
            "S": 28,
            "T": 28,
            "U": 25,
            "V": 20,
            "W": 25,
            "X": 26,
            "Y": 30,
            "Z": 30,
            "AA": 26,
            "AB": 26,
            "AC": 26,
            "AD": 26,
            "AE": 14,
            "AF": 24,
        })

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if period == "all":
            file_name = "full_cars_all.xlsx"
        else:
            file_name = f"full_cars_{start_datetime.strftime('%Y-%m-%d_%H-%M')}_{end_datetime.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    return render(request, "defects_app/export_full_cars.html")


# Процент прямого схода ППС
@login_required
def qrqc_direct_pass_api_view(request):
    day = request.GET.get("day")

    if not day:
        return JsonResponse({
            "percent": 0,
            "total_bestenevaya": 0,
            "direct_count": 0,
        })

    try:
        selected_day = datetime.strptime(day, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({
            "percent": 0,
            "total_bestenevaya": 0,
            "direct_count": 0,
        })

    total_bestenevaya = Avtomobili.objects.filter(
        data_prohoda_bestenevaya__date=selected_day
    ).count()

    direct_count = StatusAvto.objects.filter(
        status="СГП",
        data_statusa__date=selected_day
    ).exclude(
        avto__statusavto__status="СНП"
    ).values("avto_id").distinct().count()

    percent = 0
    if total_bestenevaya > 0:
        percent = round((direct_count / total_bestenevaya) * 100, 2)

    if percent > 100:
        percent = 100

    return JsonResponse({
        "percent": percent,
        "total_bestenevaya": total_bestenevaya,
        "direct_count": direct_count,
    })

@login_required
def defect_types_dashboard_view(request):
    if not can_view_reports_exports(request.user):
        return HttpResponseForbidden("У вас нет доступа к отчетам и выгрузкам.")

    response = manager_required_view(request)
    if response:
        return response

    return render(request, "defects_app/defect_types_dashboard.html")


@login_required
def defect_types_dashboard_api_view(request):
    response = manager_required_view(request)
    if response:
        return response

    period = request.GET.get("period", "today")
    date_from = request.GET.get("date_from")
    time_from = request.GET.get("time_from") or "00:00"
    date_to = request.GET.get("date_to")
    time_to = request.GET.get("time_to") or "23:59"

    now = timezone.now()

    if period == "today":
        start_datetime = datetime.combine(now.date(), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "shift":
        start_datetime = datetime.combine(now.date(), time(8, 0))
        end_datetime = datetime.combine(now.date(), time(17, 0))

    elif period == "week":
        start_datetime = datetime.combine(now.date() - timedelta(days=7), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "month":
        start_datetime = datetime.combine(now.date() - timedelta(days=30), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "custom":
        if not date_from or not date_to:
            return JsonResponse({"error": "Не выбрана дата начала или окончания."}, status=400)

        start_datetime = datetime.strptime(f"{date_from} {time_from}", "%Y-%m-%d %H:%M")
        end_datetime = datetime.strptime(f"{date_to} {time_to}", "%Y-%m-%d %H:%M")

    else:
        return JsonResponse({"error": "Некорректный период."}, status=400)

    departments = {
        "all": None,
        "production": "Производство",
        "suppliers": "Поставщики",
        "logistics": "Отдел логистики",
        "engineering": "Инженерный отдел",
        "technical": "Технологический отдел",
    }

    def build_chart_data(responsible_name=None):
        qs = Defekty.objects.filter(
            data__gte=start_datetime,
            data__lte=end_datetime,
            tip__isnull=False,
        )

        if responsible_name:
            qs = qs.filter(otvetstvennyj__nazvanie=responsible_name)

        raw_data = (
            qs.values("tip__nazvanie")
            .annotate(count=Count("id"))
            .order_by("-count", "tip__nazvanie")
        )

        total = sum(item["count"] for item in raw_data)
        cumulative = 0
        result = []

        for item in raw_data:
            cumulative += item["count"]

            pareto_percent = 0
            if total > 0:
                pareto_percent = round((cumulative / total) * 100, 2)

            result.append({
                "type": item["tip__nazvanie"],
                "count": item["count"],
                "pareto": pareto_percent,
            })

        return result

    return JsonResponse({
        "period": period,
        "start": start_datetime.strftime("%d.%m.%Y %H:%M"),
        "end": end_datetime.strftime("%d.%m.%Y %H:%M"),
        "charts": {
            "all": build_chart_data(),
            "production": build_chart_data(departments["production"]),
            "suppliers": build_chart_data(departments["suppliers"]),
            "logistics": build_chart_data(departments["logistics"]),
            "engineering": build_chart_data(departments["engineering"]),
            "technical": build_chart_data(departments["technical"]),
        }
    })

def get_defect_type_period(request):
    data = request.POST if request.method == "POST" else request.GET

    period = data.get("period", "today")
    now = timezone.now()

    if period == "today":
        start_datetime = datetime.combine(now.date(), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "shift":
        start_datetime = datetime.combine(now.date(), time(8, 0))
        end_datetime = datetime.combine(now.date(), time(17, 0))

    elif period == "week":
        start_datetime = datetime.combine(now.date() - timedelta(days=7), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "month":
        start_datetime = datetime.combine(now.date() - timedelta(days=30), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    elif period == "custom":
        date_from = data.get("date_from")
        time_from = data.get("time_from") or "00:00"
        date_to = data.get("date_to")
        time_to = data.get("time_to") or "23:59"

        start_datetime = datetime.strptime(f"{date_from} {time_from}", "%Y-%m-%d %H:%M")
        end_datetime = datetime.strptime(f"{date_to} {time_to}", "%Y-%m-%d %H:%M")

    else:
        start_datetime = datetime.combine(now.date(), time.min)
        end_datetime = datetime.combine(now.date(), time.max)

    return start_datetime, end_datetime


def build_pareto_rows(queryset):
    total = sum(item["count"] for item in queryset)

    rows = []
    cumulative = 0

    for item in queryset:
        count = item["count"]
        cumulative += count

        pareto = round((cumulative / total) * 100, 2) if total else 0

        rows.append({
            "type": item["tip__nazvanie"] or "Не указано",
            "department": item["otvetstvennyj__nazvanie"] or "Не указано",
            "count": count,
            "pareto": pareto,
        })

    return rows

def build_export_pareto_rows(queryset, include_department=True):
    total = sum(item["count"] for item in queryset)

    rows = []
    cumulative = 0

    for item in queryset:
        cumulative += item["count"]

        if total > 0:
            pareto = round(cumulative / total * 100, 2)
        else:
            pareto = 0

        row = {
            "type": item.get("tip__nazvanie") or "Не указано",
            "department": item.get("otvetstvennyj__nazvanie") or "Не указано",
            "count": item["count"],
            "pareto": pareto,
        }

        rows.append(row)

    return rows


def create_defect_type_sheet(wb, title, period_text, rows, image_data=None, include_department=True):
    ws = wb.create_sheet(title=title[:31])

    ws.merge_cells("A1:D1")
    ws["A1"] = "Отчёт по дефектам по типам"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = period_text
    ws["A2"].font = Font(size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    header_row = 4

    if include_department:
        headers = ["Тип дефекта", "Отдел", "Количество дефектов", "Парето, %"]
    else:
        headers = ["Тип дефекта", "Количество дефектов", "Парето, %"]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="B9A7CC")
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = header_row + 1

    for item in rows:
        values = [item["type"], item["department"], item["count"], item["pareto"]] if include_department else [
            item["type"], item["count"], item["pareto"]
        ]

        for col_index, value in enumerate(values, start=1):
            ws.cell(row=current_row, column=col_index).value = value

        current_row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28 if include_department else 22
    ws.column_dimensions["C"].width = 22 if include_department else 14
    ws.column_dimensions["D"].width = 14

    for row in ws.iter_rows(min_row=1, max_row=max(current_row, 8), min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(
                left=Side(style="thin", color="999999"),
                right=Side(style="thin", color="999999"),
                top=Side(style="thin", color="999999"),
                bottom=Side(style="thin", color="999999"),
            )
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    if not rows:
        ws["A5"] = "Нет данных за выбранный период"
        ws["A5"].font = Font(bold=True, color="777777")

    if image_data:
        image_data = image_data.split(",")[1]
        image_bytes = base64.b64decode(image_data)

        tmp = NamedTemporaryFile(delete=False, suffix=".png")
        tmp.write(image_bytes)
        tmp.close()

        img = ExcelImage(tmp.name)
        img.width = 900
        img.height = 430

        background_fill = PatternFill("solid", fgColor="E6E6E6")

        for row in ws.iter_rows(min_row=4, max_row=28, min_col=6, max_col=20):
            for cell in row:
                cell.fill = background_fill

        ws.add_image(img, "F4")

    return ws


@login_required
def export_defect_types_dashboard_view(request):
    response = manager_required_view(request)
    if response:
        return response

    start_datetime, end_datetime = get_defect_type_period(request)

    period_text = f"Период отчёта: {start_datetime.strftime('%d.%m.%Y %H:%M')} — {end_datetime.strftime('%d.%m.%Y %H:%M')}"
    chart_images = json.loads(request.POST.get("chart_images", "{}"))

    department_map = {
        "technical": "Технологический отдел",
        "engineering": "Инженерный отдел",
        "suppliers": "Поставщики",
        "logistics": "Отдел логистики",
        "production": "Производство",
    }

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    all_queryset = (
        Defekty.objects.filter(
            data__gte=start_datetime,
            data__lte=end_datetime,
            tip__isnull=False,
        )
        .values("tip__nazvanie", "otvetstvennyj__nazvanie")
        .annotate(count=Count("id"))
        .order_by("-count", "tip__nazvanie")
    )

    all_rows = build_export_pareto_rows(all_queryset, include_department=True)

    create_defect_type_sheet(
        wb=wb,
        title="Все отделы",
        period_text=period_text,
        rows=all_rows,
        image_data=chart_images.get("all"),
        include_department=True
    )

    for key, department_name in department_map.items():
        queryset = (
            Defekty.objects.filter(
                data__gte=start_datetime,
                data__lte=end_datetime,
                tip__isnull=False,
                otvetstvennyj__nazvanie=department_name,
            )
            .values("tip__nazvanie", "otvetstvennyj__nazvanie")
            .annotate(count=Count("id"))
            .order_by("-count", "tip__nazvanie")
        )

        rows = build_export_pareto_rows(queryset, include_department=False)

        create_defect_type_sheet(
            wb=wb,
            title=department_name,
            period_text=period_text,
            rows=rows,
            image_data=chart_images.get(key),
            include_department=False
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    file_name = (
        f"defect_types_dashboard_"
        f"{start_datetime.strftime('%Y-%m-%d_%H-%M')}_"
        f"{end_datetime.strftime('%Y-%m-%d_%H-%M')}.xlsx"
    )

    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response